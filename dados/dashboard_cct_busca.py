"""
DASHBOARD MONSTRO - BUSCA DE CONVENCOES COLETIVAS DE TRABALHO (CCT)
====================================================================
Gera planilha Excel profissional com:
  1. Busca        -> Formulario com dropdown + fórmulas INDEX/MATCH
  2. Lista        -> Tabela completa com filtros e formatação
  3. Estatísticas -> Resumo gráfico e analítico
  4. Comparativo  -> Compare 2 CCTs lado a lado
  5. Pisos        -> Apenas CCTs com pisos salariais
  6. Contribuicoes-> Analise de contribuicoes sindicais
  7. Vigencia     -> CCTs por periodo de vigencia

Atualizavel automaticamente: rode o script e o dashboard é recriado.
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
from openpyxl.comments import Comment
from datetime import datetime
import os
import re

# ============================================================================
# CONFIGURACOES
# ============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
INPUT_FILE = os.path.join(BASE_DIR, 'dashboard_cct.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, 'dashboard_cct_busca.xlsx')

# Paleta de cores profissional (dark theme / corporate)
COR_PRIMARIA = "1B4F72"       # Azul petroleo
COR_SECUNDARIA = "2E86C1"     # Azul medio
COR_ACENTO = "F39C12"         # Laranja/dourado
COR_FUNDO = "EBF5FB"          # Azul bem claro
COR_FUNDO_ESCURO = "D6EAF8"   # Azul claro
COR_BRANCO = "FFFFFF"
COR_CINZA = "F2F3F4"
COR_CINZA_ESCURO = "5D6D7E"
COR_VERDE = "27AE60"
COR_VERMELHO = "C0392B"
COR_ROXO = "8E44AD"

# Bordas
border_fina = Border(
    left=Side(style='thin', color='BDC3C7'),
    right=Side(style='thin', color='BDC3C7'),
    top=Side(style='thin', color='BDC3C7'),
    bottom=Side(style='thin', color='BDC3C7')
)
border_medium = Border(
    left=Side(style='medium', color='2E86C1'),
    right=Side(style='medium', color='2E86C1'),
    top=Side(style='medium', color='2E86C1'),
    bottom=Side(style='medium', color='2E86C1')
)
border_bottom = Border(bottom=Side(style='medium', color='2E86C1'))

# Fontes
font_titulo = Font(name='Calibri', size=20, bold=True, color=COR_PRIMARIA)
font_subtitulo = Font(name='Calibri', size=11, italic=True, color=COR_CINZA_ESCURO)
font_header = Font(name='Calibri', size=11, bold=True, color=COR_BRANCO)
font_header_dark = Font(name='Calibri', size=11, bold=True, color=COR_PRIMARIA)
font_normal = Font(name='Calibri', size=11, color="000000")
font_negrito = Font(name='Calibri', size=11, bold=True, color="000000")
font_label = Font(name='Calibri', size=11, bold=True, color=COR_PRIMARIA)
font_valor = Font(name='Calibri', size=11, color="000000")
font_destaque = Font(name='Calibri', size=14, bold=True, color=COR_ACENTO)
font_info = Font(name='Calibri', size=10, color=COR_CINZA_ESCURO)

# Fills
fill_primaria = PatternFill(start_color=COR_PRIMARIA, end_color=COR_PRIMARIA, fill_type="solid")
fill_secundaria = PatternFill(start_color=COR_SECUNDARIA, end_color=COR_SECUNDARIA, fill_type="solid")
fill_fundo = PatternFill(start_color=COR_FUNDO, end_color=COR_FUNDO, fill_type="solid")
fill_fundo_escuro = PatternFill(start_color=COR_FUNDO_ESCURO, end_color=COR_FUNDO_ESCURO, fill_type="solid")
fill_cinza = PatternFill(start_color=COR_CINZA, end_color=COR_CINZA, fill_type="solid")
fill_branco = PatternFill(start_color=COR_BRANCO, end_color=COR_BRANCO, fill_type="solid")
fill_acento = PatternFill(start_color=COR_ACENTO, end_color=COR_ACENTO, fill_type="solid")
fill_verde = PatternFill(start_color=COR_VERDE, end_color=COR_VERDE, fill_type="solid")
fill_vermelho = PatternFill(start_color=COR_VERMELHO, end_color=COR_VERMELHO, fill_type="solid")
fill_roxo = PatternFill(start_color=COR_ROXO, end_color=COR_ROXO, fill_type="solid")

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)
align_right = Alignment(horizontal='right', vertical='center', wrap_text=True)


def ajustar_larguras(ws, colunas_larguras):
    for col, largura in colunas_larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


def formatar_header(ws, row, cols, fill=None):
    if fill is None:
        fill = fill_primaria
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.font = font_header
        cell.fill = fill
        cell.alignment = align_center
        cell.border = border_fina


def linha_zebrada(ws, row, cols, par=True):
    fill = fill_cinza if par else fill_branco
    for c in cols:
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = border_fina
        cell.font = font_normal


def criar_dropdown(ws, cell_ref, formula_lista, prompt="Selecione..."):
    dv = DataValidation(type="list", formula1=formula_lista, allow_blank=False)
    dv.prompt = prompt
    dv.promptTitle = "Seleção"
    ws.add_data_validation(dv)
    dv.add(cell_ref)


# ============================================================================
# CARREGAR DADOS
# ============================================================================
print("Carregando dados de CCT...")
df = pd.read_excel(INPUT_FILE, sheet_name='Dados')

# Limpar e normalizar
for col in df.columns:
    if df[col].dtype == object:
        df[col] = df[col].astype(str).str.replace('\t', ' ').str.strip()
        df[col] = df[col].replace('nan', '')
        df[col] = df[col].replace('None', '')

# Garantir que id é numérico
df['id'] = pd.to_numeric(df['id'], errors='coerce').fillna(0).astype(int)

# Extrair codigo do sindicato dos arquivos reais na pasta convencoes
PARENT_DIR = os.path.dirname(BASE_DIR)
CONV_DIR = os.path.join(PARENT_DIR, 'convencoes')

conv_files = os.listdir(CONV_DIR) if os.path.exists(CONV_DIR) else []
conv_mapping = {}
for f in conv_files:
    m = re.match(r'^(\d+)-(?:TA-)?CCT-(.+)-\d{2}-\d{2}-\d{4}\.pdf$', f)
    if m:
        cod = int(m.group(1))
        nome_parte = m.group(2)
        chave = nome_parte[:50].upper()
        if chave not in conv_mapping:
            conv_mapping[chave] = cod

codigos = []
for nome in df['nome_arquivo']:
    nome_clean = re.sub(r'^CCT-', '', str(nome))
    nome_clean = re.sub(r'-\d{2}-\d{2}-\d{4}\.pdf$', '', nome_clean)
    chave = nome_clean[:50].upper()
    cod = conv_mapping.get(chave, 0)
    if cod == 0:
        for k, v in conv_mapping.items():
            if k in chave or chave in k:
                cod = v
                break
    codigos.append(cod)

df['cod_sindicato'] = codigos

# Criar campo de busca combinado
busca_cols = ['parte_empregados', 'parte_empregadores', 'nome_arquivo']
df['busca_completa'] = df[busca_cols].apply(
    lambda x: ' | '.join([str(v) for v in x if v and str(v).strip() and str(v) != 'nan']), axis=1
)

# Contagem de preenchimento por CCT
cols_analise = ['reajuste', 'pisos_salariais', 'contribuicao_empregados',
                'contribuicao_patronal', 'beneficios', 'jornada', 'aviso_previo', 'multa']
df['preenchimento'] = df[cols_analise].apply(lambda x: x.notna().sum(), axis=1)
df['percentual_preenchimento'] = (df['preenchimento'] / len(cols_analise) * 100).round(1)

# Categorizar por vigencia
df['tem_vigencia'] = (df['vigencia_inicio'].notna() & (df['vigencia_inicio'] != '')).astype(int)

TOTAL_CCTS = len(df)
print(f"Total de CCTs carregadas: {TOTAL_CCTS}")

# ============================================================================
# CRIAR WORKBOOK
# ============================================================================
wb = Workbook()

# ============================================================================
# ABA AUXILIAR (oculta)
# ============================================================================
ws_aux = wb.active
ws_aux.title = "_aux"
ws_aux.sheet_state = "hidden"

# Lista de nomes para dropdown (coluna A)
ws_aux.cell(row=1, column=1, value="Lista de CCTs para dropdown")
for i, nome in enumerate(df['nome_arquivo'].tolist(), start=2):
    ws_aux.cell(row=i, column=1, value=nome)

# Lista de codigos sindicato para dropdown comparativo (coluna B)
ws_aux.cell(row=1, column=2, value="Codigos Sindicato")
codigos_unicos = sorted(df['cod_sindicato'].unique())
for i, cod in enumerate(codigos_unicos, start=2):
    ws_aux.cell(row=i, column=2, value=int(cod))

# Lista de IDs (coluna C)
ws_aux.cell(row=1, column=3, value="IDs")
for i, _id in enumerate(df['id'].tolist(), start=2):
    ws_aux.cell(row=i, column=3, value=_id)

print("Aba auxiliar criada.")

# ============================================================================
# ABA 1: BUSCA (Formulario principal)
# ============================================================================
ws_busca = wb.create_sheet(title="Busca")

# Titulo
ws_busca.merge_cells('A1:H1')
c = ws_busca['A1']
c.value = "DASHBOARD DE BUSCA - CONVENÇÕES COLETIVAS DE TRABALHO"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_busca.row_dimensions[1].height = 35

ws_busca.merge_cells('A2:H2')
c = ws_busca['A2']
c.value = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}  |  Total de CCTs: {TOTAL_CCTS}"
c.font = font_subtitulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_busca.row_dimensions[2].height = 20

# Painel de busca (linhas 4-6)
ws_busca.merge_cells('A4:H4')
c = ws_busca['A4']
c.value = "SELECIONE UMA CONVENÇÃO NO DROPDOWN ABAIXO"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_secundaria
c.alignment = Alignment(horizontal='center', vertical='center')
ws_busca.row_dimensions[4].height = 28

# Dropdown de seleção
ws_busca.merge_cells('B5:G5')
c = ws_busca['B5']
c.value = df['nome_arquivo'].iloc[0]  # valor padrão
c.font = Font(name='Calibri', size=12, bold=False, color="000000")
c.fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
c.alignment = align_left
c.border = border_medium
ws_busca.row_dimensions[5].height = 30

# Dropdown
lista_ref = f"'_aux'!$A$2:$A${TOTAL_CCTS + 1}"
criar_dropdown(ws_busca, 'B5', lista_ref, "Selecione uma CCT da lista")

# Instrucoes
ws_busca['A6'] = "Dica: Clique na célula B5 acima para abrir a lista de CCTs. Os dados abaixo atualizam automaticamente."
ws_busca['A6'].font = Font(name='Calibri', size=9, italic=True, color=COR_CINZA_ESCURO)
ws_busca.merge_cells('A6:H6')

# ============================================================================
# PAINEL DE DETALHES (esquerda e direita)
# ============================================================================

# Colunas: A=Label, B=Valor, C=espaco, D=Label, E=Valor, F=espaco, G=Label, H=Valor
# Usaremos fórmulas INDEX/MATCH para buscar dinamicamente

# Determinar a linha onde cada id está na aba _aux (para MATCH)
# Vamos usar formulas que buscam na aba Dados (que vamos criar depois)
# Mas como o workbook é novo, vamos referenciar a aba "Lista" que vamos criar

row_start = 8

# --- SEÇÃO IDENTIFICAÇÃO ---
ws_busca.merge_cells(f'A{row_start}:H{row_start}')
c = ws_busca[f'A{row_start}']
c.value = "IDENTIFICAÇÃO DA CONVENÇÃO"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_primaria
c.alignment = Alignment(horizontal='left', vertical='center')
ws_busca.row_dimensions[row_start].height = 25

# Mapear colunas do df para nomes amigaveis
campos_busca = [
    ('A', 'B', 'ID:', 'id'),
    ('C', 'D', 'Código Sindicato:', 'cod_sindicato'),
    ('E', 'F', 'Nome do Arquivo:', 'nome_arquivo'),
    ('A', 'B', 'Parte Empregados:', 'parte_empregados'),
    ('C', 'D', 'CNPJ Empregados:', 'cnpj_empregados'),
    ('E', 'F', 'Parte Empregadores:', 'parte_empregadores'),
    ('A', 'B', 'CNPJ Empregadores:', 'cnpj_empregadores'),
]

# Para usar formulas INDEX/MATCH, precisamos da aba "Lista" já criada.
# Vamos criar a aba Lista primeiro, depois voltamos aqui.
# Mas podemos usar a aba Dados do arquivo original? Não, vamos criar nossa própria.
# Estratégia: criar todas as abas primeiro, depois voltar e inserir as formulas.

print("Aba Busca - estrutura base criada.")

# ============================================================================
# ABA 2: LISTA COMPLETA (tabela formatada com filtros)
# ============================================================================
ws_lista = wb.create_sheet(title="Lista")

# Titulo
ws_lista.merge_cells('A1:R1')
c = ws_lista['A1']
c.value = "LISTA COMPLETA DE CONVENÇÕES COLETIVAS DE TRABALHO"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_lista.row_dimensions[1].height = 35

ws_lista.merge_cells('A2:R2')
c = ws_lista['A2']
c.value = f"Use os filtros automáticos no cabeçalho (Ctrl+Shift+L) para buscar. Total: {TOTAL_CCTS} CCTs."
c.font = font_subtitulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_lista.row_dimensions[2].height = 20

# Headers
headers_lista = ['ID', 'Código', 'Nome do Arquivo', 'Parte Empregados', 'Parte Empregadores',
                 'CNPJ Empregados', 'CNPJ Empregadores', 'Vigência Início', 'Vigência Fim',
                 'Data-Base', 'Reajuste', 'Pisos Salariais', 'Contrib. Empregados',
                 'Contrib. Patronal', 'Benefícios', 'Jornada', 'Aviso Prévio', 'Multa',
                 'Preenchimento %']

for col, h in enumerate(headers_lista, 1):
    cell = ws_lista.cell(row=4, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_primaria
    cell.alignment = align_center
    cell.border = border_fina

# Auto-filter
ws_lista.auto_filter.ref = f"A4:S{4 + TOTAL_CCTS}"

# Dados
for i, (_, row) in enumerate(df.iterrows()):
    r = 5 + i
    is_par = (i % 2 == 0)
    fill = fill_cinza if is_par else fill_branco

    valores = [
        row['id'], row['cod_sindicato'], row['nome_arquivo'], row['parte_empregados'],
        row['parte_empregadores'], row['cnpj_empregados'], row['cnpj_empregadores'],
        row['vigencia_inicio'], row['vigencia_fim'], row['data_base'], row['reajuste'],
        row['pisos_salariais'], row['contribuicao_empregados'], row['contribuicao_patronal'],
        row['beneficios'], row['jornada'], row['aviso_previo'], row['multa'],
        row['percentual_preenchimento']
    ]

    for col, val in enumerate(valores, 1):
        cell = ws_lista.cell(row=r, column=col, value=val)
        cell.fill = fill
        cell.border = border_fina
        cell.font = font_normal
        cell.alignment = align_left if col in [3, 4, 5, 8, 9, 10, 12, 13, 14, 15, 16, 17, 18] else align_center

    # Colorir preenchimento %
    cell_pct = ws_lista.cell(row=r, column=19)
    pct = row['percentual_preenchimento']
    if pct >= 75:
        cell_pct.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    elif pct >= 50:
        cell_pct.fill = PatternFill(start_color="FCF3CF", end_color="FCF3CF", fill_type="solid")
    elif pct >= 25:
        cell_pct.fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    else:
        cell_pct.fill = PatternFill(start_color="E59866", end_color="E59866", fill_type="solid")
        cell_pct.font = Font(name='Calibri', size=11, bold=True, color=COR_BRANCO)

ajustar_larguras(ws_lista, {
    1: 6, 2: 8, 3: 55, 4: 40, 5: 40, 6: 18, 7: 18, 8: 16, 9: 16, 10: 16,
    11: 12, 12: 25, 13: 25, 14: 25, 15: 25, 16: 20, 17: 20, 18: 20, 19: 14
})

# Congelar painel
ws_lista.freeze_panes = 'A5'

print("Aba Lista criada.")

# ============================================================================
# VOLTAR A ABA BUSCA - Inserir formulas INDEX/MATCH
# ============================================================================

# A formula INDEX/MATCH busca na aba "Lista"
# O MATCH encontra a linha do nome_arquivo selecionado em B5
# O INDEX retorna o valor da coluna desejada

# Vamos usar: =INDEX(Lista!$A$5:$S$<last>, MATCH(Busca!$B$5, Lista!$C$5:$C$<last>, 0), <col>)
last_row = 4 + TOTAL_CCTS

def formula_index(col_idx):
    return f'=INDEX(Lista!$A$5:$S${last_row},MATCH(Busca!$B$5,Lista!$C$5:$C${last_row},0),{col_idx})'

# Coluna A=id(1), B=nome_arquivo(3), C=parte_empregados(4), D=cnpj_empregados(6)
# E=parte_empregadores(5), F=cnpj_empregadores(7), G=vigencia_inicio(8), H=vigencia_fim(9)
# I=data_base(10), J=reajuste(11), K=pisos(12), L=contrib_emp(13), M=contrib_pat(14)
# N=beneficios(15), O=jornada(16), P=aviso(17), Q=multa(18)

# Re-layout: vamos fazer um layout de 3 colunas de pares label/valor
row = 9

# Secao Identificacao
ws_busca.merge_cells(f'A{row}:H{row}')
c = ws_busca[f'A{row}']
c.value = "📋 IDENTIFICAÇÃO"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_primaria
c.alignment = Alignment(horizontal='left', vertical='center')
ws_busca.row_dimensions[row].height = 25

row += 1

# ID + Codigo + Nome Arquivo
ws_busca[f'A{row}'] = "ID:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca[f'B{row}'] = formula_index(1)
ws_busca[f'B{row}'].font = font_negrito
ws_busca[f'B{row}'].alignment = align_left
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom

ws_busca[f'C{row}'] = "Código Sindicato:"
ws_busca[f'C{row}'].font = font_label
ws_busca[f'C{row}'].alignment = align_right
ws_busca[f'C{row}'].fill = fill_fundo
ws_busca[f'D{row}'] = formula_index(2)
ws_busca[f'D{row}'].font = font_negrito
ws_busca[f'D{row}'].alignment = align_left
ws_busca[f'D{row}'].fill = fill_branco
ws_busca[f'D{row}'].border = border_bottom

ws_busca[f'E{row}'] = "Nome do Arquivo:"
ws_busca[f'E{row}'].font = font_label
ws_busca[f'E{row}'].alignment = align_right
ws_busca[f'E{row}'].fill = fill_fundo
ws_busca.merge_cells(f'F{row}:H{row}')
ws_busca[f'F{row}'] = formula_index(3)
ws_busca[f'F{row}'].font = Font(name='Calibri', size=10, bold=True, color="000000")
ws_busca[f'F{row}'].alignment = align_left
ws_busca[f'F{row}'].fill = fill_branco
ws_busca[f'F{row}'].border = border_bottom

row += 1

# Partes
ws_busca[f'A{row}'] = "Parte Empregados:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca.merge_cells(f'B{row}:C{row}')
ws_busca[f'B{row}'] = formula_index(4)
ws_busca[f'B{row}'].font = font_valor
ws_busca[f'B{row}'].alignment = align_left
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom

ws_busca[f'D{row}'] = "Parte Empregadores:"
ws_busca[f'D{row}'].font = font_label
ws_busca[f'D{row}'].alignment = align_right
ws_busca[f'D{row}'].fill = fill_fundo
ws_busca.merge_cells(f'E{row}:H{row}')
ws_busca[f'E{row}'] = formula_index(5)
ws_busca[f'E{row}'].font = font_valor
ws_busca[f'E{row}'].alignment = align_left
ws_busca[f'E{row}'].fill = fill_branco
ws_busca[f'E{row}'].border = border_bottom

row += 1

# CNPJs
ws_busca[f'A{row}'] = "CNPJ Empregados:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca[f'B{row}'] = formula_index(6)
ws_busca[f'B{row}'].font = font_valor
ws_busca[f'B{row}'].alignment = align_left
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom

ws_busca[f'C{row}'] = "CNPJ Empregadores:"
ws_busca[f'C{row}'].font = font_label
ws_busca[f'C{row}'].alignment = align_right
ws_busca[f'C{row}'].fill = fill_fundo
ws_busca.merge_cells(f'D{row}:H{row}')
ws_busca[f'D{row}'] = formula_index(7)
ws_busca[f'D{row}'].font = font_valor
ws_busca[f'D{row}'].alignment = align_left
ws_busca[f'D{row}'].fill = fill_branco
ws_busca[f'D{row}'].border = border_bottom

row += 2

# Secao Vigencia
ws_busca.merge_cells(f'A{row}:H{row}')
c = ws_busca[f'A{row}']
c.value = "📅 VIGÊNCIA E REAJUSTE"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_secundaria
c.alignment = Alignment(horizontal='left', vertical='center')
ws_busca.row_dimensions[row].height = 25

row += 1

ws_busca[f'A{row}'] = "Vigência Início:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca[f'B{row}'] = formula_index(8)
ws_busca[f'B{row}'].font = font_valor
ws_busca[f'B{row}'].alignment = align_left
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom

ws_busca[f'C{row}'] = "Vigência Fim:"
ws_busca[f'C{row}'].font = font_label
ws_busca[f'C{row}'].alignment = align_right
ws_busca[f'C{row}'].fill = fill_fundo
ws_busca[f'D{row}'] = formula_index(9)
ws_busca[f'D{row}'].font = font_valor
ws_busca[f'D{row}'].alignment = align_left
ws_busca[f'D{row}'].fill = fill_branco
ws_busca[f'D{row}'].border = border_bottom

ws_busca[f'E{row}'] = "Data-Base:"
ws_busca[f'E{row}'].font = font_label
ws_busca[f'E{row}'].alignment = align_right
ws_busca[f'E{row}'].fill = fill_fundo
ws_busca[f'F{row}'] = formula_index(10)
ws_busca[f'F{row}'].font = font_valor
ws_busca[f'F{row}'].alignment = align_left
ws_busca[f'F{row}'].fill = fill_branco
ws_busca[f'F{row}'].border = border_bottom

ws_busca[f'G{row}'] = "Reajuste:"
ws_busca[f'G{row}'].font = font_label
ws_busca[f'G{row}'].alignment = align_right
ws_busca[f'G{row}'].fill = fill_fundo
ws_busca[f'H{row}'] = formula_index(11)
ws_busca[f'H{row}'].font = Font(name='Calibri', size=12, bold=True, color=COR_ACENTO)
ws_busca[f'H{row}'].alignment = align_left
ws_busca[f'H{row}'].fill = PatternFill(start_color="FFF9E6", end_color="FFF9E6", fill_type="solid")
ws_busca[f'H{row}'].border = border_bottom

row += 2

# Secao Cláusulas
ws_busca.merge_cells(f'A{row}:H{row}')
c = ws_busca[f'A{row}']
c.value = "📑 CLÁUSULAS E BENEFÍCIOS"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_primaria
c.alignment = Alignment(horizontal='left', vertical='center')
ws_busca.row_dimensions[row].height = 25

row += 1

# Pisos + Contrib Empregados
ws_busca[f'A{row}'] = "Pisos Salariais:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca.merge_cells(f'B{row}:D{row}')
ws_busca[f'B{row}'] = formula_index(12)
ws_busca[f'B{row}'].font = font_valor
ws_busca[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom
ws_busca.row_dimensions[row].height = 60

ws_busca[f'E{row}'] = "Contrib. Empregados:"
ws_busca[f'E{row}'].font = font_label
ws_busca[f'E{row}'].alignment = align_right
ws_busca[f'E{row}'].fill = fill_fundo
ws_busca.merge_cells(f'F{row}:H{row}')
ws_busca[f'F{row}'] = formula_index(13)
ws_busca[f'F{row}'].font = font_valor
ws_busca[f'F{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'F{row}'].fill = fill_branco
ws_busca[f'F{row}'].border = border_bottom

row += 1

# Contrib Patronal + Beneficios
ws_busca[f'A{row}'] = "Contrib. Patronal:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca.merge_cells(f'B{row}:D{row}')
ws_busca[f'B{row}'] = formula_index(14)
ws_busca[f'B{row}'].font = font_valor
ws_busca[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom
ws_busca.row_dimensions[row].height = 50

ws_busca[f'E{row}'] = "Benefícios:"
ws_busca[f'E{row}'].font = font_label
ws_busca[f'E{row}'].alignment = align_right
ws_busca[f'E{row}'].fill = fill_fundo
ws_busca.merge_cells(f'F{row}:H{row}')
ws_busca[f'F{row}'] = formula_index(15)
ws_busca[f'F{row}'].font = font_valor
ws_busca[f'F{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'F{row}'].fill = fill_branco
ws_busca[f'F{row}'].border = border_bottom

row += 1

# Jornada + Aviso + Multa
ws_busca[f'A{row}'] = "Jornada:"
ws_busca[f'A{row}'].font = font_label
ws_busca[f'A{row}'].alignment = align_right
ws_busca[f'A{row}'].fill = fill_fundo
ws_busca.merge_cells(f'B{row}:C{row}')
ws_busca[f'B{row}'] = formula_index(16)
ws_busca[f'B{row}'].font = font_valor
ws_busca[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'B{row}'].fill = fill_branco
ws_busca[f'B{row}'].border = border_bottom
ws_busca.row_dimensions[row].height = 40

ws_busca[f'D{row}'] = "Aviso Prévio:"
ws_busca[f'D{row}'].font = font_label
ws_busca[f'D{row}'].alignment = align_right
ws_busca[f'D{row}'].fill = fill_fundo
ws_busca.merge_cells(f'E{row}:F{row}')
ws_busca[f'E{row}'] = formula_index(17)
ws_busca[f'E{row}'].font = font_valor
ws_busca[f'E{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'E{row}'].fill = fill_branco
ws_busca[f'E{row}'].border = border_bottom

ws_busca[f'G{row}'] = "Multa:"
ws_busca[f'G{row}'].font = font_label
ws_busca[f'G{row}'].alignment = align_right
ws_busca[f'G{row}'].fill = fill_fundo
ws_busca[f'H{row}'] = formula_index(18)
ws_busca[f'H{row}'].font = font_valor
ws_busca[f'H{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
ws_busca[f'H{row}'].fill = fill_branco
ws_busca[f'H{row}'].border = border_bottom

# Ajustar larguras
ajustar_larguras(ws_busca, {1: 18, 2: 22, 3: 18, 4: 22, 5: 18, 6: 22, 7: 14, 8: 22})

print("Formulas INDEX/MATCH inseridas na aba Busca.")

# ============================================================================
# ABA 3: ESTATISTICAS
# ============================================================================
ws_est = wb.create_sheet(title="Estatísticas")

# Titulo
ws_est.merge_cells('A1:H1')
c = ws_est['A1']
c.value = "ESTATÍSTICAS DAS CONVENÇÕES COLETIVAS"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_est.row_dimensions[1].height = 35

# Cards superiores
row = 3
cards = [
    ('A', 'B', 'TOTAL DE CCTs', TOTAL_CCTS, COR_PRIMARIA),
    ('C', 'D', 'COM REAJUSTE', len(df[df['reajuste'].notna() & (df['reajuste'] != '')]), COR_VERDE),
    ('E', 'F', 'COM PISOS SALARIAIS', len(df[df['pisos_salariais'].notna() & (df['pisos_salariais'] != '')]), COR_ACENTO),
    ('G', 'H', 'COM CONTRIB. EMP.', len(df[df['contribuicao_empregados'].notna() & (df['contribuicao_empregados'] != '')]), COR_SECUNDARIA),
]

for col_l, col_v, label, valor, cor in cards:
    ws_est.merge_cells(f'{col_l}{row}:{col_v}{row}')
    c = ws_est[f'{col_l}{row}']
    c.value = label
    c.font = Font(name='Calibri', size=10, bold=True, color=COR_BRANCO)
    c.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
    c.alignment = align_center
    ws_est.row_dimensions[row].height = 22

    ws_est.merge_cells(f'{col_l}{row+1}:{col_v}{row+1}')
    c = ws_est[f'{col_l}{row+1}']
    c.value = valor
    c.font = Font(name='Calibri', size=22, bold=True, color=cor)
    c.alignment = align_center
    c.fill = fill_branco
    c.border = Border(
        left=Side(style='medium', color=cor),
        right=Side(style='medium', color=cor),
        bottom=Side(style='medium', color=cor)
    )
    ws_est.row_dimensions[row+1].height = 35

row = 6

# Tabela por código de sindicato
ws_est.merge_cells(f'A{row}:H{row}')
c = ws_est[f'A{row}']
c.value = "CCTs POR CÓDIGO DE SINDICATO"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_primaria
c.alignment = Alignment(horizontal='left', vertical='center')
ws_est.row_dimensions[row].height = 25

row += 1
headers_est = ["Código Sindicato", "Qtd CCTs", "% do Total", "Com Reajuste", "Com Pisos", "Com Contrib.", "Média Preenchimento %", "Status"]
for col, h in enumerate(headers_est, 1):
    cell = ws_est.cell(row=row, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_primaria
    cell.alignment = align_center
    cell.border = border_fina

row += 1
agrupado = df.groupby('cod_sindicato').agg(
    qtd=('id', 'count'),
    reajuste=('reajuste', lambda x: (x.notna() & (x != '')).sum()),
    pisos=('pisos_salariais', lambda x: (x.notna() & (x != '')).sum()),
    contrib=('contribuicao_empregados', lambda x: (x.notna() & (x != '')).sum()),
    preenchimento=('percentual_preenchimento', 'mean')
).reset_index().sort_values('qtd', ascending=False)

for i, (_, r) in enumerate(agrupado.iterrows()):
    fill = fill_cinza if i % 2 == 0 else fill_branco
    pct = (r['qtd'] / TOTAL_CCTS) * 100
    status = "✅ Completa" if r['preenchimento'] >= 60 else ("⚠️ Regular" if r['preenchimento'] >= 30 else "❌ Básica")
    status_cor = COR_VERDE if r['preenchimento'] >= 60 else (COR_ACENTO if r['preenchimento'] >= 30 else COR_VERMELHO)

    vals = [int(r['cod_sindicato']), int(r['qtd']), f"{pct:.1f}%", int(r['reajuste']),
            int(r['pisos']), int(r['contrib']), f"{r['preenchimento']:.1f}%", status]

    for col, val in enumerate(vals, 1):
        cell = ws_est.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = border_fina
        cell.font = font_normal
        cell.alignment = align_center
        if col == 8:
            cell.font = Font(name='Calibri', size=10, bold=True, color=status_cor)

    row += 1

ajustar_larguras(ws_est, {1: 16, 2: 10, 3: 10, 4: 12, 5: 11, 6: 12, 7: 18, 8: 14})

# ============================================================================
# ABA 4: COMPARATIVO
# ============================================================================
ws_comp = wb.create_sheet(title="Comparativo")

ws_comp.merge_cells('A1:H1')
c = ws_comp['A1']
c.value = "COMPARATIVO DE CCTs - LADO A LADO"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_comp.row_dimensions[1].height = 35

ws_comp.merge_cells('A2:H2')
c = ws_comp['A2']
c.value = "Selecione duas CCTs nos dropdowns abaixo para comparar"
c.font = font_subtitulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_comp.row_dimensions[2].height = 20

# Dropdowns
ws_comp['B3'] = "CCT 1 (Esquerda):"
ws_comp['B3'].font = font_label
ws_comp['B3'].alignment = align_right
ws_comp.merge_cells('C3:E3')
ws_comp['C3'] = df['nome_arquivo'].iloc[0]
ws_comp['C3'].fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
ws_comp['C3'].border = border_medium
ws_comp['C3'].alignment = align_left
criar_dropdown(ws_comp, 'C3', lista_ref, "Selecione a primeira CCT")

ws_comp['B4'] = "CCT 2 (Direita):"
ws_comp['B4'].font = font_label
ws_comp['B4'].alignment = align_right
ws_comp.merge_cells('C4:E4')
ws_comp['C4'] = df['nome_arquivo'].iloc[1] if len(df) > 1 else df['nome_arquivo'].iloc[0]
ws_comp['C4'].fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
ws_comp['C4'].border = border_medium
ws_comp['C4'].alignment = align_left
criar_dropdown(ws_comp, 'C4', lista_ref, "Selecione a segunda CCT")

# Tabela comparativa
comp_fields = [
    ('ID', 1), ('Código Sindicato', 2), ('Parte Empregados', 4),
    ('Parte Empregadores', 5), ('CNPJ Empregados', 6), ('CNPJ Empregadores', 7),
    ('Vigência Início', 8), ('Vigência Fim', 9), ('Data-Base', 10),
    ('Reajuste', 11), ('Pisos Salariais', 12), ('Contrib. Empregados', 13),
    ('Contrib. Patronal', 14), ('Benefícios', 15), ('Jornada', 16),
    ('Aviso Prévio', 17), ('Multa', 18)
]

row = 6
ws_comp.merge_cells(f'A{row}:H{row}')
c = ws_comp[f'A{row}']
c.value = "COMPARAÇÃO DETALHADA"
c.font = Font(name='Calibri', size=12, bold=True, color=COR_BRANCO)
c.fill = fill_primaria
c.alignment = Alignment(horizontal='left', vertical='center')
ws_comp.row_dimensions[row].height = 25

row += 1
ws_comp['A'+str(row)] = "Campo"
ws_comp['A'+str(row)].font = font_header
ws_comp['A'+str(row)].fill = fill_primaria
ws_comp['A'+str(row)].alignment = align_center
ws_comp['A'+str(row)].border = border_fina

ws_comp.merge_cells(f'B{row}:D{row}')
ws_comp['B'+str(row)] = "CCT 1"
ws_comp['B'+str(row)].font = font_header
ws_comp['B'+str(row)].fill = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
ws_comp['B'+str(row)].alignment = align_center
ws_comp['B'+str(row)].border = border_fina

ws_comp.merge_cells(f'E{row}:G{row}')
ws_comp['E'+str(row)] = "CCT 2"
ws_comp['E'+str(row)].font = font_header
ws_comp['E'+str(row)].fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
ws_comp['E'+str(row)].alignment = align_center
ws_comp['E'+str(row)].border = border_fina

ws_comp['H'+str(row)] = "Igual?"
ws_comp['H'+str(row)].font = font_header
ws_comp['H'+str(row)].fill = fill_primaria
ws_comp['H'+str(row)].alignment = align_center
ws_comp['H'+str(row)].border = border_fina

row += 1

for label, col_idx in comp_fields:
    is_par = (row % 2 == 0)
    fill = fill_cinza if is_par else fill_branco

    ws_comp[f'A{row}'] = label
    ws_comp[f'A{row}'].font = font_label
    ws_comp[f'A{row}'].fill = fill_fundo
    ws_comp[f'A{row}'].alignment = align_right
    ws_comp[f'A{row}'].border = border_fina

    # CCT 1
    ws_comp.merge_cells(f'B{row}:D{row}')
    ws_comp[f'B{row}'] = f'=IFERROR(INDEX(Lista!$A$5:$S${last_row},MATCH(Comparativo!$C$3,Lista!$C$5:$C${last_row},0),{col_idx}),"")'
    ws_comp[f'B{row}'].font = font_valor
    ws_comp[f'B{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws_comp[f'B{row}'].fill = fill
    ws_comp[f'B{row}'].border = border_fina
    ws_comp.row_dimensions[row].height = max(20, 15)

    # CCT 2
    ws_comp.merge_cells(f'E{row}:G{row}')
    ws_comp[f'E{row}'] = f'=IFERROR(INDEX(Lista!$A$5:$S${last_row},MATCH(Comparativo!$C$4,Lista!$C$5:$C${last_row},0),{col_idx}),"")'
    ws_comp[f'E{row}'].font = font_valor
    ws_comp[f'E{row}'].alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws_comp[f'E{row}'].fill = fill
    ws_comp[f'E{row}'].border = border_fina

    # Igual?
    ws_comp[f'H{row}'] = f'=IF(B{row}=E{row},"✅","❌")'
    ws_comp[f'H{row}'].font = font_negrito
    ws_comp[f'H{row}'].alignment = align_center
    ws_comp[f'H{row}'].fill = fill
    ws_comp[f'H{row}'].border = border_fina

    row += 1

ajustar_larguras(ws_comp, {1: 20, 2: 20, 3: 20, 4: 20, 5: 20, 6: 20, 7: 20, 8: 10})

print("Aba Comparativo criada.")

# ============================================================================
# ABA 5: PISOS SALARIAIS
# ============================================================================
ws_pisos = wb.create_sheet(title="Pisos Salariais")

ws_pisos.merge_cells('A1:H1')
c = ws_pisos['A1']
c.value = "CCTs COM PISOS SALARIAIS"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_pisos.row_dimensions[1].height = 35

ws_pisos.merge_cells('A2:H2')
c = ws_pisos['A2']
c.value = f"Total de CCTs com pisos salariais: {len(df[df['pisos_salariais'].notna() & (df['pisos_salariais'] != '')])}"
c.font = font_subtitulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_pisos.row_dimensions[2].height = 20

headers_pisos = ['ID', 'Código', 'Nome do Arquivo', 'Parte Empregados', 'Parte Empregadores',
                 'Data-Base', 'Reajuste', 'Pisos Salariais']
row = 4
for col, h in enumerate(headers_pisos, 1):
    cell = ws_pisos.cell(row=row, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_acento
    cell.alignment = align_center
    cell.border = border_fina

df_pisos = df[df['pisos_salariais'].notna() & (df['pisos_salariais'] != '')].copy()
row = 5
for i, (_, r) in enumerate(df_pisos.iterrows()):
    fill = fill_cinza if i % 2 == 0 else fill_branco
    vals = [r['id'], r['cod_sindicato'], r['nome_arquivo'], r['parte_empregados'],
            r['parte_empregadores'], r['data_base'], r['reajuste'], r['pisos_salariais']]
    for col, val in enumerate(vals, 1):
        cell = ws_pisos.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = border_fina
        cell.font = font_normal
        cell.alignment = align_left if col in [3, 4, 5, 8] else align_center
    ws_pisos.row_dimensions[row].height = max(18, 15)
    row += 1

ajustar_larguras(ws_pisos, {1: 6, 2: 8, 3: 50, 4: 35, 5: 35, 6: 14, 7: 12, 8: 40})

print("Aba Pisos Salariais criada.")

# ============================================================================
# ABA 6: CONTRIBUICOES
# ============================================================================
ws_contrib = wb.create_sheet(title="Contribuições")

ws_contrib.merge_cells('A1:H1')
c = ws_contrib['A1']
c.value = "ANÁLISE DE CONTRIBUIÇÕES SINDICAIS"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_contrib.row_dimensions[1].height = 35

ws_contrib.merge_cells('A2:H2')
c = ws_contrib['A2']
c.value = "CCTs com contribuição de empregados e/ou patronal"
c.font = font_subtitulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_contrib.row_dimensions[2].height = 20

headers_contrib = ['ID', 'Código', 'Nome do Arquivo', 'Parte Empregados', 'Parte Empregadores',
                   'Contrib. Empregados', 'Contrib. Patronal', 'Possui Ambas?']
row = 4
for col, h in enumerate(headers_contrib, 1):
    cell = ws_contrib.cell(row=row, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_roxo
    cell.alignment = align_center
    cell.border = border_fina

df_contrib = df[(df['contribuicao_empregados'].notna() & (df['contribuicao_empregados'] != '')) |
                (df['contribuicao_patronal'].notna() & (df['contribuicao_patronal'] != ''))].copy()

row = 5
for i, (_, r) in enumerate(df_contrib.iterrows()):
    fill = fill_cinza if i % 2 == 0 else fill_branco
    tem_ambas = (r['contribuicao_empregados'] not in ['', 'nan']) and (r['contribuicao_patronal'] not in ['', 'nan'])
    vals = [r['id'], r['cod_sindicato'], r['nome_arquivo'], r['parte_empregados'],
            r['parte_empregadores'], r['contribuicao_empregados'], r['contribuicao_patronal'],
            'SIM' if tem_ambas else 'NÃO']
    for col, val in enumerate(vals, 1):
        cell = ws_contrib.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = border_fina
        cell.font = font_normal
        cell.alignment = align_left if col in [3, 4, 5, 6, 7] else align_center
        if col == 8 and tem_ambas:
            cell.font = Font(name='Calibri', size=11, bold=True, color=COR_VERDE)
            cell.fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    ws_contrib.row_dimensions[row].height = max(18, 15)
    row += 1

ajustar_larguras(ws_contrib, {1: 6, 2: 8, 3: 45, 4: 30, 5: 30, 6: 30, 7: 30, 8: 14})

print("Aba Contribuições criada.")

# ============================================================================
# ABA 7: VIGENCIA
# ============================================================================
ws_vig = wb.create_sheet(title="Vigência")

ws_vig.merge_cells('A1:H1')
c = ws_vig['A1']
c.value = "CCTs POR PERÍODO DE VIGÊNCIA"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_vig.row_dimensions[1].height = 35

ws_vig.merge_cells('A2:H2')
c = ws_vig['A2']
c.value = "Análise temporal das convenções coletivas"
c.font = font_subtitulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_vig.row_dimensions[2].height = 20

headers_vig = ['ID', 'Código', 'Nome do Arquivo', 'Parte Empregados', 'Parte Empregadores',
               'Vigência Início', 'Vigência Fim', 'Duração Estimada']
row = 4
for col, h in enumerate(headers_vig, 1):
    cell = ws_vig.cell(row=row, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_secundaria
    cell.alignment = align_center
    cell.border = border_fina

df_vig = df[df['vigencia_inicio'].notna() & (df['vigencia_inicio'] != '')].copy()
row = 5
for i, (_, r) in enumerate(df_vig.iterrows()):
    fill = fill_cinza if i % 2 == 0 else fill_branco
    vals = [r['id'], r['cod_sindicato'], r['nome_arquivo'], r['parte_empregados'],
            r['parte_empregadores'], r['vigencia_inicio'], r['vigencia_fim'], '']
    for col, val in enumerate(vals, 1):
        cell = ws_vig.cell(row=row, column=col, value=val)
        cell.fill = fill
        cell.border = border_fina
        cell.font = font_normal
        cell.alignment = align_left if col in [3, 4, 5] else align_center
    ws_vig.row_dimensions[row].height = max(18, 15)
    row += 1

ajustar_larguras(ws_vig, {1: 6, 2: 8, 3: 50, 4: 30, 5: 30, 6: 16, 7: 16, 8: 18})

print("Aba Vigência criada.")

# ============================================================================
# ABA 8: INDICE / AJUDA
# ============================================================================
ws_ajuda = wb.create_sheet(title="Índice")

ws_ajuda.merge_cells('A1:D1')
c = ws_ajuda['A1']
c.value = "📘 GUIA DE USO DO DASHBOARD"
c.font = font_titulo
c.alignment = Alignment(horizontal='center', vertical='center')
ws_ajuda.row_dimensions[1].height = 35

instrucoes = [
    ("", ""),
    ("🎯 ABA BUSCA", "Use o dropdown na célula B5 para selecionar uma CCT. Todos os campos abaixo atualizam automaticamente via fórmulas Excel."),
    ("", ""),
    ("📋 ABA LISTA", "Tabela completa com TODAS as CCTs. Use os filtros automáticos (setinhas no cabeçalho) para buscar por qualquer campo."),
    ("", ""),
    ("📊 ABA ESTATÍSTICAS", "Resumo numérico e porcentagens por código de sindicato. Identifique rapidamente quais sindicatos têm mais CCTs."),
    ("", ""),
    ("⚖️ ABA COMPARATIVO", "Selecione duas CCTs nos dropdowns e compare lado a lado. A coluna 'Igual?' mostra se os campos são idênticos."),
    ("", ""),
    ("💰 ABA PISOS SALARIAIS", "Apenas CCTs que possuem pisos salariais preenchidos. Útil para consultas rápidas de salários mínimos por categoria."),
    ("", ""),
    ("🤝 ABA CONTRIBUIÇÕES", "CCTs com contribuição sindical/negocial de empregados e/ou patronal. Identifique quais têm ambas as contribuições."),
    ("", ""),
    ("📅 ABA VIGÊNCIA", "CCTs organizadas por período de vigência. Útil para verificar quais convenções estão vigentes ou venceram."),
    ("", ""),
    ("🔄 ATUALIZAÇÃO", "Para atualizar este dashboard, rode o script Python 'dashboard_cct_busca.py'. Ele lê os dados da aba 'Dados' do arquivo 'dashboard_cct.xlsx' e recria toda a planilha."),
    ("", ""),
    ("⚠️ IMPORTANTE", "Não edite as abas '_aux' (oculta) e nem as fórmulas nas abas Busca e Comparativo, pois elas dependem da estrutura da aba Lista."),
]

row = 3
for label, texto in instrucoes:
    if label:
        ws_ajuda[f'A{row}'] = label
        ws_ajuda[f'A{row}'].font = Font(name='Calibri', size=12, bold=True, color=COR_PRIMARIA)
        ws_ajuda[f'A{row}'].alignment = align_left
    if texto:
        ws_ajuda.merge_cells(f'B{row}:D{row}')
        ws_ajuda[f'B{row}'] = texto
        ws_ajuda[f'B{row}'].font = font_normal
        ws_ajuda[f'B{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws_ajuda.row_dimensions[row].height = max(18, len(texto) // 60 * 15)
    row += 1

ajustar_larguras(ws_ajuda, {1: 25, 2: 30, 3: 30, 4: 30})

print("Aba Índice criada.")

# ============================================================================
# ORDENAR ABAS
# ============================================================================
ordem = ["Busca", "Lista", "Estatísticas", "Comparativo", "Pisos Salariais",
         "Contribuições", "Vigência", "Índice", "_aux"]
for i, nome in enumerate(ordem):
    if nome in wb.sheetnames:
        wb.move_sheet(nome, offset=-wb.sheetnames.index(nome) + i)

# Esconder aba _aux
if '_aux' in wb.sheetnames:
    wb['_aux'].sheet_state = "hidden"

# ============================================================================
# SALVAR
# ============================================================================
wb.save(OUTPUT_FILE)
print(f"\n{'='*60}")
print("DASHBOARD MONSTRO GERADO COM SUCESSO!")
print(f"{'='*60}")
print(f"Arquivo: {OUTPUT_FILE}")
print(f"\nAbas criadas ({len(wb.sheetnames)}):")
for s in wb.sheetnames:
    print(f"  • {s}")
print(f"\nTotal de CCTs processadas: {TOTAL_CCTS}")
print(f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print(f"\nPara atualizar, execute: python dashboard_cct_busca.py")
