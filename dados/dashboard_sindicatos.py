"""
Dashboard de Sindicatos x Empresas
Gera planilha Excel caprichada mostrando:
  - Por Sindicato: quais empresas pertencem a cada sindicato/codigo
  - Por Empresa: quais sindicatos cada empresa possui
  - Resumo: estatisticas gerais
Considera codsindicato e codsindicato2 (empresa pode ter ate 2 sindicatos)
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import os

# ============================================================================
# CONFIGURACOES
# ============================================================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# O script fica dentro de /dados/; os dados-fonte ficam no diretorio pai
PARENT_DIR = os.path.dirname(BASE_DIR)
SINDICATOS_FILE = os.path.join(PARENT_DIR, 'sindicatosistema.xlsx')
EMPRESAS_FILE = os.path.join(PARENT_DIR, 'empresasindicato.xlsx')
OUTPUT_FILE = os.path.join(BASE_DIR, 'dashboard_sindicatos.xlsx')

# Cores do tema
COR_TITULO_SINDICATO = "1F4E78"      # Azul escuro
COR_TITULO_EMPRESA = "2E7D32"        # Verde escuro
COR_HEADER = "D9E1F2"                # Azul claro
COR_HEADER_EMP = "C8E6C9"            # Verde claro
COR_LINHA_PAR = "F2F2F2"             # Cinza bem claro
COR_LINHA_IMPAR = "FFFFFF"           # Branco
COR_DESTAQUE = "FFF3CD"              # Amarelo claro

# Estilos
font_titulo = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
font_header = Font(name='Calibri', size=11, bold=True, color="000000")
font_normal = Font(name='Calibri', size=11, color="000000")
font_negrito = Font(name='Calibri', size=11, bold=True, color="000000")
font_resumo = Font(name='Calibri', size=12, bold=True, color="1F4E78")

fill_titulo_sind = PatternFill(start_color=COR_TITULO_SINDICATO, end_color=COR_TITULO_SINDICATO, fill_type="solid")
fill_titulo_emp = PatternFill(start_color=COR_TITULO_EMPRESA, end_color=COR_TITULO_EMPRESA, fill_type="solid")
fill_header = PatternFill(start_color=COR_HEADER, end_color=COR_HEADER, fill_type="solid")
fill_header_emp = PatternFill(start_color=COR_HEADER_EMP, end_color=COR_HEADER_EMP, fill_type="solid")
fill_par = PatternFill(start_color=COR_LINHA_PAR, end_color=COR_LINHA_PAR, fill_type="solid")
fill_impar = PatternFill(start_color=COR_LINHA_IMPAR, end_color=COR_LINHA_IMPAR, fill_type="solid")
fill_destaque = PatternFill(start_color=COR_DESTAQUE, end_color=COR_DESTAQUE, fill_type="solid")

border_fina = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC')
)
border_box = Border(
    left=Side(style='medium', color='000000'),
    right=Side(style='medium', color='000000'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='medium', color='000000')
)

align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
align_left = Alignment(horizontal='left', vertical='center', wrap_text=True)


def ajustar_larguras(ws, colunas_larguras):
    for col, largura in colunas_larguras.items():
        ws.column_dimensions[get_column_letter(col)].width = largura


def aplicar_bordas_faixa(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_fina


# ============================================================================
# CARREGAR DADOS
# ============================================================================
print("Carregando dados...")
df_sind = pd.read_excel(SINDICATOS_FILE)
df_emp = pd.read_excel(EMPRESAS_FILE)

# Limpar espacos e tabs
df_sind['sindicato'] = df_sind['sindicato'].astype(str).str.replace('\t', ' ').str.strip()
df_sind['codigo'] = df_sind['codigo'].astype(int)
df_emp['empresa'] = df_emp['empresa'].astype(str).str.strip()
df_emp['codempresa'] = df_emp['codempresa'].astype(int)

# Dicionario codigo -> nome do sindicato
sindicato_map = dict(zip(df_sind['codigo'], df_sind['sindicato']))
# Garantir que codigo 0 tenha descricao padrao
if 0 not in sindicato_map:
    sindicato_map[0] = "Sem Enquadramento Sindical Definido"

# ============================================================================
# PREPARAR RELACOES
# ============================================================================
print("Preparando relacoes...")

# Lista de relacoes: (cod_sindicato, nome_sindicato, cod_empresa, nome_empresa)
relacoes = []

for _, row in df_emp.iterrows():
    cod_emp = row['codempresa']
    nome_emp = row['empresa']

    # Sindicato 1
    cod_s1 = row['codsindicato']
    if pd.notna(cod_s1):
        cod_s1 = int(cod_s1)
        nome_s1 = sindicato_map.get(cod_s1, "SINDICATO NAO ENCONTRADO")
        relacoes.append((cod_s1, nome_s1, cod_emp, nome_emp))

    # Sindicato 2
    cod_s2 = row['codsindicato2']
    if pd.notna(cod_s2):
        cod_s2 = int(cod_s2)
        nome_s2 = sindicato_map.get(cod_s2, "SINDICATO NAO ENCONTRADO")
        relacoes.append((cod_s2, nome_s2, cod_emp, nome_emp))

df_rel = pd.DataFrame(relacoes, columns=['cod_sindicato', 'nome_sindicato', 'cod_empresa', 'nome_empresa'])

# Agrupar por sindicato
sindicatos_ordenados = sorted(df_rel['cod_sindicato'].unique())

# Agrupar por empresa para aba "Por Empresa"
empresas_info = []
for _, row in df_emp.iterrows():
    cod_emp = row['codempresa']
    nome_emp = row['empresa']
    sinds = []
    if pd.notna(row['codsindicato']):
        c = int(row['codsindicato'])
        sinds.append(f"{c} - {sindicato_map.get(c, 'N/A')}")
    if pd.notna(row['codsindicato2']):
        c = int(row['codsindicato2'])
        sinds.append(f"{c} - {sindicato_map.get(c, 'N/A')}")
    empresas_info.append({
        'cod_empresa': cod_emp,
        'nome_empresa': nome_emp,
        'sindicatos': sinds,
        'qtd': len(sinds)
    })

# ============================================================================
# CRIAR WORKBOOK
# ============================================================================
print("Criando planilha Excel...")
wb = Workbook()

# ============================================================================
# ABA 1: RESUMO
# ============================================================================
ws_resumo = wb.active
ws_resumo.title = "Resumo"

# Titulo
ws_resumo.merge_cells('A1:D1')
cell = ws_resumo['A1']
cell.value = "DASHBOARD SINDICATOS x EMPRESAS"
cell.font = Font(name='Calibri', size=18, bold=True, color="1F4E78")
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_resumo.row_dimensions[1].height = 30

# Subtitulo
ws_resumo.merge_cells('A2:D2')
cell = ws_resumo['A2']
cell.value = f"Total de Sindicatos: {len(sindicato_map)}  |  Total de Empresas: {len(df_emp)}  |  Relacoes ativas: {len(df_rel)}"
cell.font = Font(name='Calibri', size=11, italic=True, color="555555")
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_resumo.row_dimensions[2].height = 20

# Espaco
ws_resumo.row_dimensions[3].height = 5

# Tabela resumo por sindicato
headers = ["Codigo Sindicato", "Nome do Sindicato", "Qtd Empresas", "% do Total"]
for col, h in enumerate(headers, 1):
    cell = ws_resumo.cell(row=4, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = align_center
    cell.border = border_fina

row_idx = 5
total_empresas = len(df_emp)
for cod_s in sindicatos_ordenados:
    nome_s = sindicato_map.get(cod_s, "N/A")
    qtd = len(df_rel[df_rel['cod_sindicato'] == cod_s])
    pct = (qtd / total_empresas) * 100 if total_empresas > 0 else 0

    ws_resumo.cell(row=row_idx, column=1, value=cod_s).alignment = align_center
    ws_resumo.cell(row=row_idx, column=2, value=nome_s).alignment = align_left
    ws_resumo.cell(row=row_idx, column=3, value=qtd).alignment = align_center
    ws_resumo.cell(row=row_idx, column=4, value=f"{pct:.1f}%").alignment = align_center

    # Cor zebrada
    fill = fill_par if row_idx % 2 == 0 else fill_impar
    for col in range(1, 5):
        ws_resumo.cell(row=row_idx, column=col).fill = fill
        ws_resumo.cell(row=row_idx, column=col).border = border_fina
        ws_resumo.cell(row=row_idx, column=col).font = font_normal

    row_idx += 1

# Total
ws_resumo.cell(row=row_idx, column=1, value="").border = border_fina
ws_resumo.cell(row=row_idx, column=2, value="TOTAL").font = font_negrito
ws_resumo.cell(row=row_idx, column=2).alignment = align_center
ws_resumo.cell(row=row_idx, column=3, value=len(df_rel)).font = font_negrito
ws_resumo.cell(row=row_idx, column=3).alignment = align_center
ws_resumo.cell(row=row_idx, column=4, value="100.0%").font = font_negrito
ws_resumo.cell(row=row_idx, column=4).alignment = align_center
for col in range(1, 5):
    ws_resumo.cell(row=row_idx, column=col).fill = fill_destaque
    ws_resumo.cell(row=row_idx, column=col).border = border_fina

ajustar_larguras(ws_resumo, {1: 18, 2: 70, 3: 15, 4: 12})

# ============================================================================
# ABA 2: POR SINDICATO
# ============================================================================
ws_sind = wb.create_sheet(title="Por Sindicato")

row_idx = 1
for idx, cod_s in enumerate(sindicatos_ordenados):
    nome_s = sindicato_map.get(cod_s, "N/A")
    empresas_sind = df_rel[df_rel['cod_sindicato'] == cod_s][['cod_empresa', 'nome_empresa']].sort_values('cod_empresa')
    qtd = len(empresas_sind)

    # Titulo do sindicato
    ws_sind.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    cell = ws_sind.cell(row=row_idx, column=1, value=f"{cod_s} - {nome_s}  ({qtd} empresa{'s' if qtd != 1 else ''})")
    cell.font = font_titulo
    cell.fill = fill_titulo_sind
    cell.alignment = align_left
    ws_sind.row_dimensions[row_idx].height = 25

    row_idx += 1

    # Headers
    ws_sind.cell(row=row_idx, column=1, value="Cod Empresa").font = font_header
    ws_sind.cell(row=row_idx, column=1).fill = fill_header
    ws_sind.cell(row=row_idx, column=1).alignment = align_center
    ws_sind.cell(row=row_idx, column=1).border = border_fina

    ws_sind.cell(row=row_idx, column=2, value="Nome da Empresa").font = font_header
    ws_sind.cell(row=row_idx, column=2).fill = fill_header
    ws_sind.cell(row=row_idx, column=2).alignment = align_left
    ws_sind.cell(row=row_idx, column=2).border = border_fina

    ws_sind.cell(row=row_idx, column=3, value="Situacao").font = font_header
    ws_sind.cell(row=row_idx, column=3).fill = fill_header
    ws_sind.cell(row=row_idx, column=3).alignment = align_center
    ws_sind.cell(row=row_idx, column=3).border = border_fina

    row_idx += 1

    # Dados
    if qtd == 0:
        ws_sind.cell(row=row_idx, column=1, value="-")
        ws_sind.cell(row=row_idx, column=2, value="Nenhuma empresa vinculada")
        ws_sind.cell(row=row_idx, column=3, value="-")
        for col in range(1, 4):
            ws_sind.cell(row=row_idx, column=col).border = border_fina
            ws_sind.cell(row=row_idx, column=col).font = font_normal
        row_idx += 1
    else:
        for i, (_, er) in enumerate(empresas_sind.iterrows()):
            fill = fill_par if i % 2 == 0 else fill_impar
            ws_sind.cell(row=row_idx, column=1, value=er['cod_empresa']).alignment = align_center
            ws_sind.cell(row=row_idx, column=2, value=er['nome_empresa']).alignment = align_left
            ws_sind.cell(row=row_idx, column=3, value="Vinculada").alignment = align_center
            for col in range(1, 4):
                c = ws_sind.cell(row=row_idx, column=col)
                c.border = border_fina
                c.fill = fill
                c.font = font_normal
            row_idx += 1

    # Espaco entre sindicatos
    row_idx += 1

ajustar_larguras(ws_sind, {1: 15, 2: 60, 3: 15})

# ============================================================================
# ABA 3: POR EMPRESA
# ============================================================================
ws_emp = wb.create_sheet(title="Por Empresa")

# Titulo geral
ws_emp.merge_cells('A1:D1')
cell = ws_emp['A1']
cell.value = "RELACAO DE EMPRESAS E SEUS SINDICATOS"
cell.font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
cell.fill = fill_titulo_emp
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_emp.row_dimensions[1].height = 28

# Headers
headers_emp = ["Cod Empresa", "Nome da Empresa", "Sindicato(s) Vinculado(s)", "Qtd Sindicatos"]
row_idx = 3
for col, h in enumerate(headers_emp, 1):
    cell = ws_emp.cell(row=row_idx, column=col, value=h)
    cell.font = font_header
    cell.fill = fill_header_emp
    cell.alignment = align_center
    cell.border = border_fina

row_idx = 4
for i, info in enumerate(sorted(empresas_info, key=lambda x: x['cod_empresa'])):
    fill = fill_par if i % 2 == 0 else fill_impar

    ws_emp.cell(row=row_idx, column=1, value=info['cod_empresa']).alignment = align_center
    ws_emp.cell(row=row_idx, column=2, value=info['nome_empresa']).alignment = align_left
    ws_emp.cell(row=row_idx, column=3, value="\n".join(info['sindicatos'])).alignment = align_left
    ws_emp.cell(row=row_idx, column=4, value=info['qtd']).alignment = align_center

    for col in range(1, 5):
        c = ws_emp.cell(row=row_idx, column=col)
        c.border = border_fina
        c.fill = fill
        c.font = font_normal

    # Altura da linha baseada em quantos sindicatos
    ws_emp.row_dimensions[row_idx].height = max(18, 16 * info['qtd'])

    row_idx += 1

ajustar_larguras(ws_emp, {1: 15, 2: 45, 3: 80, 4: 16})

# ============================================================================
# ABA 4: EMPRESAS MULTI-SINDICATO
# ============================================================================
ws_multi = wb.create_sheet(title="Multi-Sindicato")

multi = [e for e in empresas_info if e['qtd'] > 1]

ws_multi.merge_cells('A1:D1')
cell = ws_multi['A1']
cell.value = f"EMPRESAS COM MAIS DE 1 SINDICATO ({len(multi)} empresas)"
cell.font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
cell.fill = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
cell.alignment = Alignment(horizontal='center', vertical='center')
ws_multi.row_dimensions[1].height = 28

headers_multi = ["Cod Empresa", "Nome da Empresa", "Sindicato(s) Vinculado(s)", "Qtd"]
row_idx = 3
for col, h in enumerate(headers_multi, 1):
    cell = ws_multi.cell(row=row_idx, column=col, value=h)
    cell.font = font_header
    cell.fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    cell.alignment = align_center
    cell.border = border_fina

row_idx = 4
for i, info in enumerate(sorted(multi, key=lambda x: x['cod_empresa'])):
    fill = fill_par if i % 2 == 0 else fill_impar
    ws_multi.cell(row=row_idx, column=1, value=info['cod_empresa']).alignment = align_center
    ws_multi.cell(row=row_idx, column=2, value=info['nome_empresa']).alignment = align_left
    ws_multi.cell(row=row_idx, column=3, value="\n".join(info['sindicatos'])).alignment = align_left
    ws_multi.cell(row=row_idx, column=4, value=info['qtd']).alignment = align_center
    for col in range(1, 5):
        c = ws_multi.cell(row=row_idx, column=col)
        c.border = border_fina
        c.fill = fill
        c.font = font_normal
    ws_multi.row_dimensions[row_idx].height = max(18, 16 * info['qtd'])
    row_idx += 1

ajustar_larguras(ws_multi, {1: 15, 2: 45, 3: 80, 4: 12})

# ============================================================================
# SALVAR
# ============================================================================
wb.save(OUTPUT_FILE)
print(f"\nDashboard gerado com sucesso!")
print(f"Arquivo: {OUTPUT_FILE}")
print(f"\nAbas criadas:")
print("  1. Resumo         - Estatisticas gerais por sindicato")
print("  2. Por Sindicato  - Empresas agrupadas por sindicato")
print("  3. Por Empresa    - Lista completa de empresas com seus sindicatos")
print("  4. Multi-Sindicato - Apenas empresas com 2 ou mais sindicatos")
