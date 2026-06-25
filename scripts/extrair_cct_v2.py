#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrator de Convenções Coletivas de Trabalho (CCT) - VERSÃO REFINADA
Extrai:
- Partes (sindicatos)
- Data base, vigência início/fim
- Contribuição sindical/negocial dos empregados (desconto)
- Contribuição sindical patronal (se houver)
- Pisos salariais por função (todos)
- Percentual de reajuste
- Benefícios com descrição
- Jornada, aviso prévio, multa (se houver algo específico)
"""

import os
import re
import glob
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
import platform

def wsl_path(win_path):
    if 'microsoft' in platform.release().lower() or 'WSL' in platform.release().upper():
        if win_path.startswith('C:\\\\'):
            return '/mnt/c' + win_path[2:].replace('\\\\', '/')
        elif win_path.startswith('C:'):
            return '/mnt/c' + win_path[2:].replace('\\', '/')
    return win_path

PASTA_PDFS_WIN = r"C:\Script\mediadorcct\convencoes"
ARQUIVO_SAIDA_WIN = r"C:\Script\mediadorcct\dados\dashboard_cct.xlsx"

PASTA_PDFS = wsl_path(PASTA_PDFS_WIN)
ARQUIVO_SAIDA = wsl_path(ARQUIVO_SAIDA_WIN)

# =============================================================================
# FUNÇÕES DE EXTRAÇÃO
# =============================================================================

def extrair_texto_pdf(caminho_pdf):
    texto_completo = []
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    texto_completo.append(text)
    except Exception as e:
        print(f"Erro ao ler {caminho_pdf}: {e}")
    return "\n".join(texto_completo)


def extrair_partes(texto):
    """
    Extrai partes (empregados/empregadores) de forma mais robusta.
    Funciona para CCTs e ACTs (acordos coletivos), incluindo termos aditivos.
    """
    linhas = texto.split('\n')
    
    # Encontrar todas as linhas com CNPJ
    linhas_com_cnpj = []
    for i, linha in enumerate(linhas):
        if re.search(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', linha):
            linhas_com_cnpj.append(i)
    
    if len(linhas_com_cnpj) < 2:
        # Fallback para o método antigo
        return extrair_partes_fallback(texto)
    
    # Ignorar CNPJs que aparecem em contextos que não são partes
    linhas_com_cnpj_validas = []
    for idx in linhas_com_cnpj:
        linha = linhas[idx]
        if re.search(r'Processo\s+n°?º?\s*:', linha, re.IGNORECASE):
            continue
        if re.search(r'Registro\s+n°?º?\s*:', linha, re.IGNORECASE):
            continue
        if 'http://www3.mte.gov.br' in linha:
            continue
        linhas_com_cnpj_validas.append(idx)
    
    if len(linhas_com_cnpj_validas) < 2:
        return extrair_partes_fallback(texto)
    
    # Encontrar o índice do separador "E" entre os dois primeiros CNPJs válidos
    idx_e = -1
    for i in range(linhas_com_cnpj_validas[0], linhas_com_cnpj_validas[1]):
        if linhas[i].strip() in ('E', 'E;', 'E,'):
            idx_e = i
            break
    
    # Se não achou "E", tenta procurar entre o primeiro e segundo CNPJ válido
    if idx_e == -1:
        for i in range(linhas_com_cnpj_validas[0] + 1, linhas_com_cnpj_validas[1]):
            if len(linhas[i].strip()) <= 2 and linhas[i].strip():
                idx_e = i
                break
    
    def extrair_nome_completo(linhas, cnpj_idx, limite_idx):
        """Extrai o nome completo retrocedendo a partir da linha do CNPJ."""
        nome_linhas = []
        
        # A linha do CNPJ pode ter parte do nome também
        linha_cnpj = linhas[cnpj_idx]
        parte_antes_cnpj = re.split(r'CNPJ', linha_cnpj, flags=re.IGNORECASE)[0]
        parte_antes_cnpj = parte_antes_cnpj.strip().rstrip(',;')
        if parte_antes_cnpj and len(parte_antes_cnpj) > 3:
            nome_linhas.insert(0, parte_antes_cnpj)
        
        # Retroceder para pegar o nome completo
        for i in range(cnpj_idx - 1, limite_idx, -1):
            linha = linhas[i].strip()
            if not linha:
                continue
            if linha.lower() in ('e', 'e;', 'e,'):
                break
            if 'celebram' in linha.lower() and len(linha) < 30:
                break
            if 'termos aditivo' in linha.lower():
                break
            if 'confira a autenticidade' in linha.lower():
                break
            if linha.startswith('NÚMERO') or linha.startswith('DATA'):
                break
            if re.search(r'Processo\s+n°', linha, re.IGNORECASE):
                break
            if re.search(r'Registro\s+n°', linha, re.IGNORECASE):
                break
            if i <= 2:
                break
            nome_linhas.insert(0, linha)
        
        nome = ' '.join(nome_linhas)
        nome = re.sub(r'\s+', ' ', nome).strip(' ,;')
        return nome
    
    limite = linhas_com_cnpj_validas[0] if idx_e == -1 else idx_e
    
    parte_empregados = extrair_nome_completo(linhas, linhas_com_cnpj_validas[0], -1)
    cnpj_empregados = extrair_cnpj(linhas[linhas_com_cnpj_validas[0]])
    
    parte_empregadores = extrair_nome_completo(linhas, linhas_com_cnpj_validas[1], limite)
    cnpj_empregadores = extrair_cnpj(linhas[linhas_com_cnpj_validas[1]])
    
    # Se tem mais de 2 CNPJs válidos, pode haver uma terceira parte
    if len(linhas_com_cnpj_validas) >= 3:
        tem_e_2_3 = False
        for i in range(linhas_com_cnpj_validas[1], linhas_com_cnpj_validas[2]):
            if linhas[i].strip() in ('E', 'E;', 'E,'):
                tem_e_2_3 = True
                break
        
        if tem_e_2_3:
            parte_empregadores = extrair_nome_completo(linhas, linhas_com_cnpj_validas[1], limite)
            cnpj_empregadores = extrair_cnpj(linhas[linhas_com_cnpj_validas[1]])
    
    return {
        'parte_empregados': limpar_nome(parte_empregados),
        'parte_empregadores': limpar_nome(parte_empregadores),
        'cnpj_empregados': cnpj_empregados,
        'cnpj_empregadores': cnpj_empregadores,
    }


def extrair_partes_fallback(texto):
    """Método antigo de fallback."""
    linhas = texto.split('\n')
    idx_e = -1
    for i, linha in enumerate(linhas):
        if linha.strip() in ('E', 'E;', 'E,'):
            idx_e = i
            break

    def extrair_nome_sindicato(linhas, start_idx, max_end_idx):
        nome_partes = []
        for i in range(start_idx, max_end_idx):
            linha = linhas[i].strip()
            if not linha:
                continue
            if 'celebram' in linha.lower():
                break
            if i > start_idx and ('SIND' in linha.upper() or 'FEDERACAO' in linha.upper()):
                break
            if linha in ('E', 'E;', 'E,'):
                break
            if 'CNPJ' in linha.upper():
                parte_antes = re.split(r'CNPJ', linha, flags=re.IGNORECASE)[0]
                if parte_antes.strip():
                    nome_partes.append(parte_antes.strip().rstrip(',;'))
                break
            if 'representado' in linha.lower():
                break
            nome_partes.append(linha)
        return ' '.join(nome_partes)

    parte_empregados = ""
    parte_empregadores = ""
    cnpj_empregados = ""
    cnpj_empregadores = ""

    if idx_e != -1:
        for i in range(idx_e + 1, len(linhas)):
            linha = linhas[i].strip()
            if 'SIND' in linha.upper() or 'FEDERACAO' in linha.upper():
                fim = min(i + 10, len(linhas))
                parte_empregadores = extrair_nome_sindicato(linhas, i, fim)
                for j in range(i, fim):
                    if 'CNPJ' in linhas[j].upper():
                        cnpj_empregadores = extrair_cnpj(linhas[j])
                break

        for i in range(0, idx_e):
            linha = linhas[i].strip()
            if 'SIND' in linha.upper() or 'FEDERACAO' in linha.upper():
                parte_empregados = extrair_nome_sindicato(linhas, i, idx_e)
                for j in range(i, idx_e):
                    if 'CNPJ' in linhas[j].upper():
                        cnpj_empregados = extrair_cnpj(linhas[j])
                break
    else:
        match = re.search(
            r'(SIND[^\n]+|FEDERACAO[^\n]+).*?\n\s*E\s*\n\s*(SIND[^\n]+|FEDERACAO[^\n]+)',
            texto, re.DOTALL | re.IGNORECASE
        )
        if match:
            parte_empregados = match.group(1).strip()
            parte_empregadores = match.group(2).strip()

    return {
        'parte_empregados': limpar_nome(parte_empregados),
        'parte_empregadores': limpar_nome(parte_empregadores),
        'cnpj_empregados': cnpj_empregados,
        'cnpj_empregadores': cnpj_empregadores,
    }


def extrair_cnpj(linha):
    match = re.search(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', linha)
    return match.group(0) if match else ""


def limpar_nome(nome):
    nome = nome.strip()
    if not nome:
        return ""
    # Cortar em certos padrões
    cortes = [
        r'CNPJ\s*n?\.', r'CNPJ\s*n?\s*\d', r'CNPJ\s*$',
        r'neste ato representado', r'representado\(a\)',
        r'Presidente,', r'Vice-Presidente,',
        r'Sócio,', r'Sócia,',
        r'Diretor,', r'Diretora,',
        r'Sr\(a\)\.', r'Sr\.', r'Sra\.', r'celebram',
    ]
    for padrao in cortes:
        nome = re.split(padrao, nome, flags=re.IGNORECASE)[0]
    
    nome = nome.strip(' ,;')
    nome = nome.replace('\n', ' ')
    nome = re.sub(r'\s+', ' ', nome)
    nome = nome.strip(' ,;')
    return nome


def extrair_vigencia(texto):
    vigencia_inicio = ""
    vigencia_fim = ""
    data_base = ""

    match = re.search(
        r'(\d{1,2}º?\s+de\s+\w+\s+de\s+\d{4})\s+a\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})',
        texto, re.IGNORECASE
    )
    if match:
        vigencia_inicio = match.group(1)
        vigencia_fim = match.group(2)

    match_db = re.search(
        r'data-base[^\d]*(\d{1,2}º?\s+de\s+\w+\s+de\s+\d{4}|\d{2}/\d{4})',
        texto, re.IGNORECASE
    )
    if match_db:
        data_base = match_db.group(1)
    else:
        match_db2 = re.search(r'data.base\s+.*?(\d{1,2}/\d{4})', texto, re.IGNORECASE)
        if match_db2:
            data_base = match_db2.group(1)

    return vigencia_inicio, vigencia_fim, data_base


def extrair_reajuste(texto):
    reajuste = ""
    match = re.search(
        r'reajust[ae]\w*.*?\s+(\d{1,2},\d{1,2})\s*%',
        texto, re.IGNORECASE
    )
    if match:
        reajuste = match.group(1) + "%"
    else:
        match2 = re.search(
            r'(\d{1,2},\d{1,2})\s*%(\s*\([^)]*\))',
            texto, re.IGNORECASE
        )
        if match2:
            reajuste = match2.group(1) + "%"
    return reajuste


def extrair_pisos_salariais(texto):
    """Extrai todos os pisos salariais por função."""
    pisos = []
    
    # Procurar seção de pisos salariais
    match_secao = re.search(
        r'(CLÁUSULA\s+\w+\s*-\s*(?:PISO\s+SALARIAL|PISOS\s+SALARIAIS)).*?(?=CLÁUSULA\s+\w+\s*-|$)',
        texto, re.IGNORECASE | re.DOTALL
    )
    
    if not match_secao:
        match_secao = re.search(
            r'(PISO\s+SALARIAL|PISOS\s+SALARIAIS).*?(?=CLÁUSULA\s+\w+\s*-|$)',
            texto, re.IGNORECASE | re.DOTALL
        )
    
    if match_secao:
        secao = match_secao.group(0)
        
        # Normalizar o texto - juntar linhas quebradas
        linhas = secao.split('\n')
        texto_normalizado = []
        linha_atual = ""
        for linha in linhas:
            if re.match(r'^([a-z])\)\s*', linha.strip(), re.IGNORECASE):
                if linha_atual:
                    texto_normalizado.append(linha_atual)
                linha_atual = linha
            elif linha.strip() and not linha.strip().startswith('PARAGR'):
                if linha_atual:
                    linha_atual += " " + linha.strip()
                else:
                    texto_normalizado.append(linha)
            else:
                if linha_atual:
                    texto_normalizado.append(linha_atual)
                    linha_atual = ""
        if linha_atual:
            texto_normalizado.append(linha_atual)
        
        # Procurar por padrões
        for linha in texto_normalizado:
            match = re.match(r'^([a-z])\)\s*(.*?)\s*(?:R\$|RS)\s*([\d\.,]+)', linha, re.IGNORECASE)
            if match:
                letra = match.group(1).upper()
                funcao = match.group(2).strip()
                valor = match.group(3)
                # Limpar
                funcao = re.sub(r'\s*(?:a partir de|de)\s*\dº.*$', '', funcao, flags=re.IGNORECASE).strip()
                funcao = re.sub(r':\s*$', '', funcao).strip()
                if funcao and len(funcao) > 5:
                    pisos.append(f"{letra}) {funcao}: R$ {valor}")
        
        # Se não encontrou por letra, tentar padrão mais genérico
        if not pisos:
            matches = re.findall(
                r'(Aos\s+[^\n]*?)(?:R\$|RS)\s*([\d\.,]+)',
                secao, re.IGNORECASE
            )
            for funcao, valor in matches:
                funcao = funcao.strip().replace('\n', ' ').strip(' ,;')
                funcao = re.sub(r'\(.*\)', '', funcao).strip()
                if len(funcao) > 10 and 'empregado' in funcao.lower():
                    pisos.append(f"{funcao}: R$ {valor}")
    
    # Se não encontrou na seção, procurar no texto todo
    if not pisos:
        matches = re.findall(
            r'([a-z])\)\s*(Aos[^\n]*?)(?:R\$|RS)\s*([\d\.,]+)',
            texto, re.IGNORECASE
        )
        for letra, funcao, valor in matches:
            funcao = funcao.strip().replace('\n', ' ').strip(' ,;')
            funcao = re.sub(r'\(.*\)', '', funcao).strip()
            if 'piso' in funcao.lower() or 'salário' in funcao.lower() or 'empregado' in funcao.lower():
                pisos.append(f"{letra.upper()}) {funcao}: R$ {valor}")
    
    # Padrão alternativo: piso salarial de R$ X para função Y
    if not pisos:
        matches = re.findall(
            r'piso\s+salarial.*?R\$\s*([\d\.,]+).*?(?:para|de|aos)\s+([^\n;\.]+)',
            texto, re.IGNORECASE
        )
        for valor, funcao in matches:
            funcao = funcao.strip().replace('\n', ' ').strip(' ,;')
            pisos.append(f"Piso: {funcao}: R$ {valor}")
    
    return "\n".join(pisos) if pisos else ""


def extrair_contribuicao_empregados(texto):
    """Extrai contribuição sindical/negocial dos empregados."""
    contribuicao = ""
    
    match = re.search(
        r'CONTRIBUIÇÃO\s+(?:SINDICAL|NEGOCIAL).*?(?=CLÁUSULA|CAPÍTULO|$)',
        texto, re.IGNORECASE | re.DOTALL
    )
    if match:
        secao = match.group(0)
        pct_match = re.search(r'(\d+[,.]?\d*)\s*%', secao)
        if pct_match:
            pct = pct_match.group(1) + "%"
            desc = secao.replace('\n', ' ').strip()
            if len(desc) > 300:
                desc = desc[:300] + "..."
            contribuicao = f"{pct} - {desc}"
        else:
            contribuicao = secao.replace('\n', ' ').strip()[:300]
    else:
        match2 = re.search(
            r'contribuição\s+(?:sindical|negocial|assistencial).*?\d+[,.\d]*\s*%.*?\n',
            texto, re.IGNORECASE
        )
        if match2:
            contribuicao = match2.group(0).strip()
    
    return contribuicao


def extrair_contribuicao_patronal(texto):
    """Extrai contribuição sindical patronal, se houver."""
    contribuicao = ""
    
    match = re.search(
        r'contribuição\s+patronal.*?(?=CLÁUSULA|CAPÍTULO|$)',
        texto, re.IGNORECASE | re.DOTALL
    )
    if match:
        contribuicao = match.group(0).replace('\n', ' ').strip()[:300]
    else:
        match2 = re.search(
            r'(contribuição\s+da\s+empresa|contribuição\s+empresarial|contribuição\s+do\s+empregador).*?(?=CLÁUSULA|CAPÍTULO|\n\n|$)',
            texto, re.IGNORECASE | re.DOTALL
        )
        if match2:
            contribuicao = match2.group(0).replace('\n', ' ').strip()[:300]
    
    return contribuicao


def extrair_beneficios_detalhados(texto):
    """Extrai benefícios com descrição do que diz cada um."""
    beneficios = []
    
    padroes_beneficios = {
        'Auxílio Transporte / Vale-transporte': [
            r'Auxílio Transporte.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'vale.transporte.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'vale.combustível.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Refeição / Alimentação (VA/VR)': [
            r'vale.refeição.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'vale.alimentação.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'auxílio\s+refeição.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'refeição\s+fornecida.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Auxílio Creche': [
            r'Auxílio Creche.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'creche.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Plano de Saúde / Assistência Médica': [
            r'plano\s+de\s+saúde.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'convênio\s+médico.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'assistência\s+médica.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Plano Odontológico': [
            r'odontológico.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'plano\s+odontológico.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Seguro de Vida': [
            r'seguro\s+de\s+vida.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Bolsa / Estudo': [
            r'bolsa\s+de\s+estudo.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'estudantes.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'vestibulandos.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Cesta Básica': [
            r'cesta\s+básica.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'PPR / Participação nos Lucros': [
            r'participação\s+nos\s+lucros.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'PPR.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'PLR.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Diárias / Ressarcimento': [
            r'diárias.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'ressarcimento\s+de\s+despesas.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Adicional de Periculosidade': [
            r'adicional\s+de\s+periculosidade.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Adicional de Insalubridade': [
            r'adicional\s+de\s+insalubridade.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Adicional Noturno': [
            r'adicional\s+noturno.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Horas Extras': [
            r'hora\s+extra.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'horas\s+extraordinárias.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Descanso Semanal Remunerado': [
            r'repouso\s+semanal.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'descanso\s+semanal.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'Abono de Faltas': [
            r'abono\s+de\s+faltas.*?(?=CLÁUSULA|CAPÍTULO|$)',
            r'abonar.*falta.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
        'FGTS': [
            r'FGTS.*?(?=CLÁUSULA|CAPÍTULO|$)',
        ],
    }
    
    for beneficio, padroes in padroes_beneficios.items():
        for padrao in padroes:
            match = re.search(padrao, texto, re.IGNORECASE | re.DOTALL)
            if match:
                descricao = match.group(0).replace('\n', ' ').strip()
                if len(descricao) > 400:
                    descricao = descricao[:400] + "..."
                beneficios.append(f"**{beneficio}**:\n{descricao}")
                break
    
    return "\n\n".join(beneficios) if beneficios else "Não especificado"


def extrair_jornada(texto):
    """Extrai informações sobre jornada, se houver algo específico além do padrão."""
    jornada = ""
    
    match = re.search(
        r'jornada\s+de\s+trabalho.*?\d{2,3}\s+horas.*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    if match:
        jornada = match.group(0).strip()
    
    match2 = re.search(
        r'(banco\s+de\s+horas|compensação\s+de\s+jornada|regime\s+de\s+plantão).*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    if match2:
        if jornada:
            jornada += " | " + match2.group(0).strip()
        else:
            jornada = match2.group(0).strip()
    
    return jornada


def extrair_aviso_previo(texto):
    """Extrai informações sobre aviso prévio, se houver algo específico."""
    aviso = ""
    
    match = re.search(
        r'aviso\s+prévio.*?\d{1,2}\s+dias.*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    if match:
        aviso = match.group(0).strip()
    
    match2 = re.search(
        r'aviso\s+prévio\s+indenizado.*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    if match2:
        if aviso:
            aviso += " | " + match2.group(0).strip()
        else:
            aviso = match2.group(0).strip()
    
    return aviso


def extrair_multa(texto):
    """Extrai informações sobre multa, se houver algo específico."""
    multa = ""
    
    match = re.search(
        r'multa\s+.*?\d+\s*(?:salário|piso|SM|reais).*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    if match:
        multa = match.group(0).strip()
    
    match2 = re.search(
        r'penalidade\w*.*?\d+.*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    if match2:
        if multa:
            multa += " | " + match2.group(0).strip()
        else:
            multa = match2.group(0).strip()
    
    return multa


# =============================================================================
# PROCESSAMENTO PRINCIPAL
# =============================================================================

def processar_pdf(caminho_pdf):
    nome_arquivo = os.path.basename(caminho_pdf)
    print(f"Processando: {nome_arquivo}")

    texto = extrair_texto_pdf(caminho_pdf)
    if not texto:
        return None

    partes = extrair_partes(texto)
    vig_inicio, vig_fim, data_base = extrair_vigencia(texto)
    reajuste = extrair_reajuste(texto)
    pisos = extrair_pisos_salariais(texto)
    contrib_empregados = extrair_contribuicao_empregados(texto)
    contrib_patronal = extrair_contribuicao_patronal(texto)
    beneficios = extrair_beneficios_detalhados(texto)
    jornada = extrair_jornada(texto)
    aviso = extrair_aviso_previo(texto)
    multa = extrair_multa(texto)

    dados = {
        'nome_arquivo': nome_arquivo,
        'parte_empregados': partes['parte_empregados'],
        'parte_empregadores': partes['parte_empregadores'],
        'cnpj_empregados': partes['cnpj_empregados'],
        'cnpj_empregadores': partes['cnpj_empregadores'],
        'vigencia_inicio': vig_inicio,
        'vigencia_fim': vig_fim,
        'data_base': data_base,
        'reajuste': reajuste,
        'pisos_salariais': pisos,
        'contribuicao_empregados': contrib_empregados,
        'contribuicao_patronal': contrib_patronal,
        'beneficios': beneficios,
        'jornada': jornada,
        'aviso_previo': aviso,
        'multa': multa,
    }

    return dados


def processar_todos_pdfs():
    pasta = PASTA_PDFS
    pdfs = glob.glob(os.path.join(pasta, "*.pdf")) + glob.glob(os.path.join(pasta, "*.PDF"))

    if not pdfs:
        print(f"Nenhum PDF encontrado em {pasta}")
        return []

    resultados = []
    for pdf in pdfs:
        dados = processar_pdf(pdf)
        if dados:
            resultados.append(dados)

    return resultados


# =============================================================================
# CRIAR EXCEL
# =============================================================================

def criar_excel(resultados):
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    df = pd.DataFrame(resultados)
    df = df.sort_values('nome_arquivo').reset_index(drop=True)
    df['id'] = df.index + 1

    colunas = [
        'id', 'nome_arquivo',
        'parte_empregados', 'parte_empregadores',
        'cnpj_empregados', 'cnpj_empregadores',
        'vigencia_inicio', 'vigencia_fim', 'data_base',
        'reajuste', 'pisos_salariais',
        'contribuicao_empregados', 'contribuicao_patronal',
        'beneficios',
        'jornada', 'aviso_previo', 'multa',
    ]
    df = df[[c for c in colunas if c in df.columns]]

    wb = Workbook()

    # === ABA 1: DADOS ===
    ws_dados = wb.active
    ws_dados.title = "Dados"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF", size=10)
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    for col in ws_dados.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        ws_dados.column_dimensions[column].width = adjusted_width

    ws_dados.freeze_panes = 'A2'
    ws_dados.auto_filter.ref = ws_dados.dimensions

    # === ABA 2: DASHBOARD ===
    ws_dash = wb.create_sheet("Dashboard")

    ws_dash['A1'] = "DASHBOARD - CONVENÇÕES COLETIVAS DE TRABALHO"
    ws_dash['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_dash['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.merge_cells('A1:F1')
    ws_dash.row_dimensions[1].height = 30

    ws_dash['A2'] = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_dash['A2'].font = Font(italic=True, color="666666")
    ws_dash.merge_cells('A2:F2')

    ws_dash['A4'] = "Selecione uma convenção no dropdown abaixo para visualizar os detalhes:"
    ws_dash['A4'].font = Font(bold=True, size=11)
    ws_dash.merge_cells('A4:F4')

    ws_dash['A5'] = "Convenção:"
    ws_dash['A5'].font = Font(bold=True)
    ws_dash['A5'].alignment = Alignment(horizontal="right", vertical="center")

    opcoes = [f"{row['id']} - {row['nome_arquivo']}" for _, row in df.iterrows()]
    col_oculta = 'H'
    for i, opcao in enumerate(opcoes, 1):
        ws_dash[f'{col_oculta}{i}'] = opcao

    dv = DataValidation(
        type="list",
        formula1=f"${col_oculta}$1:${col_oculta}${len(opcoes)}",
        allow_blank=True
    )
    dv.error = 'Selecione uma convenção da lista'
    dv.errorTitle = 'Seleção inválida'
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash['B5'])

    ws_dash['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.row_dimensions[5].height = 25
    ws_dash.column_dimensions['B'].width = 80

    campos = [
        ('A7', 'Empregados (Parte):', 'B7', 'parte_empregados'),
        ('A8', 'Empregadores (Parte):', 'B8', 'parte_empregadores'),
        ('A9', 'CNPJ Empregados:', 'B9', 'cnpj_empregados'),
        ('A10', 'CNPJ Empregadores:', 'B10', 'cnpj_empregadores'),
        ('A11', 'Vigência:', 'B11', 'vigencia_fim'),
        ('A12', 'Data-Base:', 'B12', 'data_base'),
        ('A13', 'Reajuste:', 'B13', 'reajuste'),
        ('A14', 'Pisos Salariais:', 'B14', 'pisos_salariais'),
        ('A15', 'Contribuição Empregados:', 'B15', 'contribuicao_empregados'),
        ('A16', 'Contribuição Patronal:', 'B16', 'contribuicao_patronal'),
        ('A17', 'Benefícios:', 'B17', 'beneficios'),
        ('A18', 'Jornada:', 'B18', 'jornada'),
        ('A19', 'Aviso Prévio:', 'B19', 'aviso_previo'),
        ('A20', 'Multa:', 'B20', 'multa'),
    ]

    for label_cell, label_text, value_cell, col_name in campos:
        ws_dash[label_cell] = label_text
        ws_dash[label_cell].font = Font(bold=True, size=10)
        ws_dash[label_cell].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        ws_dash[label_cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = ws_dados.cell(row=1, column=col_idx).column_letter
        formula = f'=IF(B5="","",VLOOKUP(VALUE(LEFT(B5,FIND(" -",B5)-1)),Dados!A:{col_letter},{col_idx},FALSE))'

        ws_dash[value_cell] = formula
        ws_dash[value_cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_dash[value_cell].font = Font(size=10)
        ws_dash[value_cell].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    ws_dash.column_dimensions['B'].width = 80
    ws_dash.column_dimensions['A'].width = 28

    for row in range(7, 21):
        ws_dash.row_dimensions[row].height = 50

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for label_cell, _, value_cell, _ in campos:
        ws_dash[label_cell].border = thin_border

    ws_dash.column_dimensions['H'].hidden = True

    ws_dash['A22'] = "Instruções:"
    ws_dash['A22'].font = Font(bold=True, size=10)
    ws_dash.merge_cells('A22:F22')
    ws_dash['A23'] = "• Para atualizar os dados, execute: python extrair_cct_v2.py"
    ws_dash['A23'].font = Font(size=9, color="666666")
    ws_dash.merge_cells('A23:F23')
    ws_dash['A24'] = "• Os dados são extraídos automaticamente dos PDFs da pasta 'convencoes'"
    ws_dash['A24'].font = Font(size=9, color="666666")
    ws_dash.merge_cells('A24:F24')

    wb.save(ARQUIVO_SAIDA)
    print(f"\nPlanilha salva em: {ARQUIVO_SAIDA}")
    print(f"Total de convenções processadas: {len(resultados)}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("EXTRATOR DE CONVENÇÕES COLETIVAS - V2 REFINADO")
    print("=" * 60)
    print(f"Pasta de origem: {PASTA_PDFS}")
    print(f"Arquivo de saída: {ARQUIVO_SAIDA}")
    print("=" * 60)

    resultados = processar_todos_pdfs()

    if resultados:
        criar_excel(resultados)
    else:
        print("Nenhum PDF processado com sucesso.")

    print("\nConcluído!")
