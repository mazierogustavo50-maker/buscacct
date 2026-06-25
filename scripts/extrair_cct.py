#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extrator de Convenções Coletivas de Trabalho (CCT)
Lê PDFs de CCTs e gera planilha Excel com:
- Partes (sindicatos)
- Vigência, data-base, reajuste, piso
- Contribuição sindical/negocial
- Benefícios
- Resumo sintético
- Dashboard para seleção de convenção
"""

import os
import re
import glob
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

# =============================================================================
# CONFIGURAÇÃO
# =============================================================================
# Detectar se estamos no WSL e converter caminho
import platform

def wsl_path(win_path):
    """Converte caminho Windows para WSL se necessário."""
    if 'microsoft' in platform.release().lower() or 'WSL' in platform.release().upper():
        if win_path.startswith('C:\\\\'):
            return '/mnt/c' + win_path[2:].replace('\\\\', '/')
        elif win_path.startswith('C:'):
            return '/mnt/c' + win_path[2:].replace('\\', '/')
    return win_path

PASTA_PDFS_WIN = r"C:\Script\mediadorcct\convencoes"
ARQUIVO_SAIDA_WIN = r"C:\Script\mediadorcct\dados\dashboard_cct.xlsx"

PASTA_PDFS = wsl_path(PASTA_PDFS_WIN)
ARQUIVO_SAIDA = wsl_path(ARQUIVO_SAIDA_WIN)

# =============================================================================
# FUNÇÕES DE EXTRAÇÃO
# =============================================================================

def extrair_texto_pdf(caminho_pdf):
    """Extrai todo o texto de um PDF."""
    texto_completo = []
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text()
                if text:
                    texto_completo.append(text)
    except Exception as e:
        print(f"Erro ao ler {caminho_pdf}: {e}")
    return "\n".join(texto_completo)


def extrair_partes(texto):
    """
    Extrai as partes da convenção:
    - Parte empregados: primeiro sindicato listado (primeiro antes do 'E')
    - Parte empregadores: primeiro sindicato DEPOIS do 'E'
    """
    linhas = texto.split('\n')

    # Encontrar a linha com "E" isolado
    idx_e = -1
    for i, linha in enumerate(linhas):
        if linha.strip() == 'E' or linha.strip() == 'E;' or linha.strip() == 'E,':
            idx_e = i
            break

    def extrair_nome_sindicato(linhas, start_idx, max_end_idx):
        """Extrai nome do sindicato concatenando linhas até CNPJ/representado."""
        nome_partes = []
        for i in range(start_idx, max_end_idx):
            linha = linhas[i].strip()
            if not linha:
                continue
            # Parar imediatamente se encontrar celebram ou outro sindicato
            if 'celebram' in linha.lower():
                break
            if i > start_idx and ('SIND' in linha.upper() or 'FEDERACAO' in linha.upper()):
                break
            if linha == 'E' or linha == 'E;' or linha == 'E,':
                break
            # Se a linha tem CNPJ, pegar apenas a parte antes do CNPJ
            if 'CNPJ' in linha.upper():
                parte_antes = re.split(r'CNPJ', linha, flags=re.IGNORECASE)[0]
                if parte_antes.strip():
                    nome_partes.append(parte_antes.strip().rstrip(',;'))
                break
            # Se a linha tem representado, parar
            if 'representado' in linha.lower():
                break
            nome_partes.append(linha)
        return ' '.join(nome_partes)

    parte_empregados = ""
    parte_empregadores = ""
    cnpj_empregados = ""
    cnpj_empregadores = ""
    rep_empregados = ""
    rep_empregadores = ""

    if idx_e != -1:
        # Parte empregadores: primeiro sindicato DEPOIS do E
        for i in range(idx_e + 1, len(linhas)):
            linha = linhas[i].strip()
            if 'SIND' in linha.upper() or 'FEDERACAO' in linha.upper():
                fim = min(i + 10, len(linhas))
                parte_empregadores = extrair_nome_sindicato(linhas, i, fim)
                for j in range(i, fim):
                    if 'CNPJ' in linhas[j].upper():
                        cnpj_empregadores = extrair_cnpj(linhas[j])
                    if 'representado' in linhas[j].lower() and not rep_empregadores:
                        rep_empregadores = extrair_representante(linhas[j])
                break

        # Parte empregados: primeiro sindicato ANTES do E
        # Vamos procurar do início até o E
        for i in range(0, idx_e):
            linha = linhas[i].strip()
            if 'SIND' in linha.upper() or 'FEDERACAO' in linha.upper():
                # Verificar se não é o empregador (deve estar antes do E)
                parte_empregados = extrair_nome_sindicato(linhas, i, idx_e)
                for j in range(i, idx_e):
                    if 'CNPJ' in linhas[j].upper():
                        cnpj_empregados = extrair_cnpj(linhas[j])
                    if 'representado' in linhas[j].lower() and not rep_empregados:
                        rep_empregados = extrair_representante(linhas[j])
                break
    else:
        # Fallback: tentar encontrar com regex
        match = re.search(
            r'(SIND[^\n]+|FEDERACAO[^\n]+).*?\n\s*E\s*\n\s*(SIND[^\n]+|FEDERACAO[^\n]+)',
            texto, re.DOTALL | re.IGNORECASE
        )
        if match:
            parte_empregados = match.group(1).strip()
            parte_empregadores = match.group(2).strip()

    return {
        'parte_empregados': limpar_nome(parte_empregados),
        'parte_empregadores': limpar_nome(parte_empregadores),
        'cnpj_empregados': cnpj_empregados,
        'cnpj_empregadores': cnpj_empregadores,
        'representante_empregados': rep_empregados,
        'representante_empregadores': rep_empregadores,
    }


def extrair_cnpj(linha):
    """Extrai CNPJ de uma linha."""
    match = re.search(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', linha)
    if match:
        return match.group(0)
    return ""


def extrair_representante(linha):
    """Extrai nome do representante."""
    match = re.search(r'Sr\(a\)\.\s*([^;]+)', linha)
    if match:
        return match.group(1).strip()
    return ""


def limpar_nome(nome):
    """Limpa nome do sindicato, remove CNPJ, representante, etc."""
    nome = nome.strip()
    # Remove CNPJ e tudo depois
    nome = re.split(r'CNPJ\s*n\.\s*\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}', nome)[0]
    nome = re.split(r'CNPJ', nome, flags=re.IGNORECASE)[0]
    # Remove representante
    nome = re.split(r'neste ato representado', nome, flags=re.IGNORECASE)[0]
    nome = re.split(r'representado\(a\)', nome, flags=re.IGNORECASE)[0]
    nome = re.split(r'celebram', nome, flags=re.IGNORECASE)[0]
    nome = re.split(r' Presidente,', nome, flags=re.IGNORECASE)[0]
    nome = re.split(r' Vice-Presidente,', nome, flags=re.IGNORECASE)[0]
    # Remove "E;" no final
    nome = nome.replace('E;', '').replace('E,', '').strip(' ;,')
    # Remove quebras de linha
    nome = nome.replace('\n', ' ').strip()
    return nome


def extrair_vigencia(texto):
    """Extrai vigência e data-base."""
    vigencia_inicio = ""
    vigencia_fim = ""
    data_base = ""

    # Padrão: "01º de junho de 2025 a 31 de maio de 2026"
    match = re.search(
        r'(\d{1,2}º?\s+de\s+\w+\s+de\s+\d{4})\s+a\s+(\d{1,2}\s+de\s+\w+\s+de\s+\d{4})',
        texto, re.IGNORECASE
    )
    if match:
        vigencia_inicio = match.group(1)
        vigencia_fim = match.group(2)

    # Data-base
    match_db = re.search(
        r'data-base[^\d]*(\d{1,2}º?\s+de\s+\w+\s+de\s+\d{4}|\d{2}/\d{4})',
        texto, re.IGNORECASE
    )
    if match_db:
        data_base = match_db.group(1)
    else:
        match_db2 = re.search(r'data.base\s+.*?(\d{1,2}/\d{4})', texto, re.IGNORECASE)
        if match_db2:
            data_base = match_db2.group(1)

    return vigencia_inicio, vigencia_fim, data_base


def extrair_reajuste(texto):
    """Extrai percentual de reajuste."""
    reajuste = ""
    # Padrão: "reajustados em ... 6,20%"
    match = re.search(
        r'reajust[ae]\w*.*?\s+(\d{1,2},\d{1,2})\s*%',
        texto, re.IGNORECASE
    )
    if match:
        reajuste = match.group(1) + "%"
    else:
        match2 = re.search(
            r'(\d{1,2},\d{1,2})\s*%(\s*\([^)]*\))',
            texto, re.IGNORECASE
        )
        if match2:
            reajuste = match2.group(1) + "%"
    return reajuste


def extrair_piso(texto):
    """Extrai piso salarial."""
    piso = ""
    match = re.search(
        r'piso\s+salarial.*?R\$\s*([\d\.]+,\d{2})',
        texto, re.IGNORECASE
    )
    if match:
        piso = "R$ " + match.group(1)
    else:
        # Procurar "R$ X.XXX,XX" próximo de piso
        match2 = re.search(
            r'piso.*?(?:\n|.){0,100}R\$\s*([\d\.]+,\d{2})',
            texto, re.IGNORECASE
        )
        if match2:
            piso = "R$ " + match2.group(1)
    return piso


def extrair_contribuicao(texto):
    """Extrai informações sobre contribuição sindical/negocial."""
    contribuicao = ""
    # Procurar cláusula de contribuição
    match = re.search(
        r'CONTRIBUIÇÃO\s+(?:SINDICAL|NEGOCIAL).*?(?=CLÁUSULA|CAPÍTULO|$)',
        texto, re.IGNORECASE | re.DOTALL
    )
    if match:
        contribuicao = match.group(0).replace('\n', ' ').strip()
        # Limitar tamanho
        if len(contribuicao) > 500:
            contribuicao = contribuicao[:500] + "..."
    else:
        # Procurar menção a contribuição
        match2 = re.search(
            r'contribuição\s+\w+.*?\d+[,.\d]*\s*%.*?\n',
            texto, re.IGNORECASE
        )
        if match2:
            contribuicao = match2.group(0).strip()
    return contribuicao


def extrair_beneficios(texto):
    """Extrai benefícios mencionados na CCT."""
    beneficios = []
    keywords = {
        'Auxílio Transporte / Vale-transporte': r'auxílio\s+transporte|vale.transporte|vale.combustível',
        'Refeição/Alimentação (VA/VR)': r'vale.refeição|vale.alimentação|auxílio\s+refeição|refeição\s+fornecida',
        'Auxílio Creche': r'auxílio\s+creche|creche',
        'Plano de Saúde': r'plano\s+de\s+saúde|convênio\s+médico|assistência\s+médica',
        'Plano Odontológico': r'odontológico|plano\s+odontológico',
        'Seguro de Vida': r'seguro\s+de\s+vida',
        'Bolsa/Estudo': r'bolsa\s+de\s+estudo|estudantes|vestibulandos',
        'Cesta Básica': r'cesta\s+básica',
        'PPR/Participação nos Lucros': r'participação\s+nos\s+lucros|PPR|PLR',
        'Diárias': r'diárias|ressarcimento\s+de\s+despesas',
        'Adicional de Periculosidade': r'adicional\s+de\s+periculosidade',
        'Adicional de Insalubridade': r'adicional\s+de\s+insalubridade',
        'Adicional Noturno': r'adicional\s+noturno',
        'Horas Extras': r'hora\s+extra|horas\s+extraordinárias',
        'Descanso Semanal Remunerado': r'repouso\s+semanal|descanso\s+semanal',
        'Abono de Faltas': r'abono\s+de\s+faltas|abonar.*falta',
        'FGTS': r'FGTS',
    }

    for beneficio, regex in keywords.items():
        if re.search(regex, texto, re.IGNORECASE):
            beneficios.append(beneficio)

    return "; ".join(beneficios) if beneficios else "Não especificado"


def extrair_outros_dados(texto):
    """Extrai outros dados relevantes."""
    dados = {}

    # Multa
    match = re.search(
        r'multa\s+.*?\d+\s*(?:salário|piso|SM).*?(?=\n|CLÁUSULA)',
        texto, re.IGNORECASE
    )
    dados['multa'] = match.group(0).strip() if match else ""

    # Jornada
    match = re.search(
        r'jornada\s+de\s+trabalho.*?\d{2,3}\s+horas',
        texto, re.IGNORECASE
    )
    dados['jornada'] = match.group(0).strip() if match else ""

    # Aviso prévio
    match = re.search(
        r'aviso\s+prévio.*?\d{1,2}\s+dias',
        texto, re.IGNORECASE
    )
    dados['aviso_previo'] = match.group(0).strip() if match else ""

    return dados


def gerar_resumo(texto, dados):
    """Gera resumo sintético da CCT."""
    resumo = []

    # Reajuste
    if dados.get('reajuste'):
        resumo.append(f"Reajuste: {dados['reajuste']}")

    # Piso
    if dados.get('piso_salarial'):
        resumo.append(f"Piso: {dados['piso_salarial']}")

    # Vigência
    if dados.get('vigencia_inicio') and dados.get('vigencia_fim'):
        resumo.append(f"Vigência: {dados['vigencia_inicio']} a {dados['vigencia_fim']}")

    # Benefícios
    if dados.get('beneficios'):
        resumo.append(f"Benefícios: {dados['beneficios']}")

    # Contribuição
    if dados.get('contribuicao'):
        contrib_resumo = dados['contribuicao'][:100]
        resumo.append(f"Contribuição: {contrib_resumo}...")

    return " | ".join(resumo)


# =============================================================================
# PROCESSAMENTO PRINCIPAL
# =============================================================================

def processar_pdf(caminho_pdf):
    """Processa um único PDF e retorna dicionário com dados."""
    nome_arquivo = os.path.basename(caminho_pdf)
    print(f"Processando: {nome_arquivo}")

    texto = extrair_texto_pdf(caminho_pdf)
    if not texto:
        return None

    # Extrair partes
    partes = extrair_partes(texto)

    # Extrair vigência
    vig_inicio, vig_fim, data_base = extrair_vigencia(texto)

    # Extrair reajuste
    reajuste = extrair_reajuste(texto)

    # Extrair piso
    piso = extrair_piso(texto)

    # Extrair contribuição
    contribuicao = extrair_contribuicao(texto)

    # Extrair benefícios
    beneficios = extrair_beneficios(texto)

    # Outros dados
    outros = extrair_outros_dados(texto)

    dados = {
        'nome_arquivo': nome_arquivo,
        'parte_empregados': partes['parte_empregados'],
        'parte_empregadores': partes['parte_empregadores'],
        'cnpj_empregados': partes['cnpj_empregados'],
        'cnpj_empregadores': partes['cnpj_empregadores'],
        'representante_empregados': partes['representante_empregados'],
        'representante_empregadores': partes['representante_empregadores'],
        'vigencia_inicio': vig_inicio,
        'vigencia_fim': vig_fim,
        'data_base': data_base,
        'reajuste': reajuste,
        'piso_salarial': piso,
        'contribuicao_sindical': contribuicao,
        'beneficios': beneficios,
        'multa': outros['multa'],
        'jornada': outros['jornada'],
        'aviso_previo': outros['aviso_previo'],
    }

    dados['resumo_sintetico'] = gerar_resumo(texto, dados)

    return dados


def processar_todos_pdfs():
    """Processa todos os PDFs da pasta."""
    pasta = PASTA_PDFS
    pdfs = glob.glob(os.path.join(pasta, "*.pdf")) + glob.glob(os.path.join(pasta, "*.PDF"))

    if not pdfs:
        print(f"Nenhum PDF encontrado em {pasta}")
        return []

    resultados = []
    for pdf in pdfs:
        dados = processar_pdf(pdf)
        if dados:
            resultados.append(dados)

    return resultados


# =============================================================================
# CRIAR EXCEL COM DASHBOARD
# =============================================================================

def criar_excel(resultados):
    """Cria planilha Excel com dados e dashboard."""
    if not resultados:
        print("Nenhum resultado para salvar.")
        return

    df = pd.DataFrame(resultados)

    # Ordenar por nome do arquivo
    df = df.sort_values('nome_arquivo').reset_index(drop=True)

    # Criar coluna de identificação
    df['id'] = df.index + 1

    # Reordenar colunas
    colunas = [
        'id', 'nome_arquivo',
        'parte_empregados', 'parte_empregadores',
        'cnpj_empregados', 'cnpj_empregadores',
        'representante_empregados', 'representante_empregadores',
        'vigencia_inicio', 'vigencia_fim', 'data_base',
        'reajuste', 'piso_salarial',
        'contribuicao_sindical',
        'beneficios',
        'multa', 'jornada', 'aviso_previo',
        'resumo_sintetico'
    ]
    df = df[[c for c in colunas if c in df.columns]]

    # Criar workbook
    wb = Workbook()

    # === ABA 1: DADOS ===
    ws_dados = wb.active
    ws_dados.title = "Dados"

    # Escrever dados
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws_dados.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

    # Ajustar largura das colunas
    for col in ws_dados.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws_dados.column_dimensions[column].width = adjusted_width

    # Congelar painel
    ws_dados.freeze_panes = 'A2'

    # Filtro
    ws_dados.auto_filter.ref = ws_dados.dimensions

    # === ABA 2: DASHBOARD ===
    ws_dash = wb.create_sheet("Dashboard")

    # Título
    ws_dash['A1'] = "DASHBOARD - CONVENÇÕES COLETIVAS DE TRABALHO"
    ws_dash['A1'].font = Font(size=16, bold=True, color="1F4E78")
    ws_dash['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.merge_cells('A1:F1')
    ws_dash.row_dimensions[1].height = 30

    # Data de atualização
    ws_dash['A2'] = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_dash['A2'].font = Font(italic=True, color="666666")
    ws_dash.merge_cells('A2:F2')

    # Instruções
    ws_dash['A4'] = "Selecione uma convenção no dropdown abaixo para visualizar os detalhes:"
    ws_dash['A4'].font = Font(bold=True, size=11)
    ws_dash.merge_cells('A4:F4')

    # Dropdown para selecionar convenção
    ws_dash['A5'] = "Convenção:"
    ws_dash['A5'].font = Font(bold=True)
    ws_dash['A5'].alignment = Alignment(horizontal="right", vertical="center")

    # Criar lista de opções para o dropdown
    opcoes = [f"{row['id']} - {row['nome_arquivo']}" for _, row in df.iterrows()]
    # Inserir opções em uma coluna oculta
    col_oculta = 'H'
    for i, opcao in enumerate(opcoes, 1):
        ws_dash[f'{col_oculta}{i}'] = opcao

    # Criar validação de dados
    dv = DataValidation(
        type="list",
        formula1=f"${col_oculta}$1:${col_oculta}${len(opcoes)}",
        allow_blank=True
    )
    dv.error = 'Selecione uma convenção da lista'
    dv.errorTitle = 'Seleção inválida'
    ws_dash.add_data_validation(dv)
    dv.add(ws_dash['B5'])

    ws_dash['B5'].alignment = Alignment(horizontal="left", vertical="center")
    ws_dash.row_dimensions[5].height = 25
    ws_dash.column_dimensions['B'].width = 80

    # Campos de detalhamento
    campos = [
        ('A7', 'Empregados (Parte):', 'B7', 'parte_empregados'),
        ('A8', 'Empregadores (Parte):', 'B8', 'parte_empregadores'),
        ('A9', 'CNPJ Empregados:', 'B9', 'cnpj_empregados'),
        ('A10', 'CNPJ Empregadores:', 'B10', 'cnpj_empregadores'),
        ('A11', 'Representante Empregados:', 'B11', 'representante_empregados'),
        ('A12', 'Representante Empregadores:', 'B12', 'representante_empregadores'),
        ('A13', 'Vigência:', 'B13', 'vigencia_fim'),
        ('A14', 'Data-Base:', 'B14', 'data_base'),
        ('A15', 'Reajuste:', 'B15', 'reajuste'),
        ('A16', 'Piso Salarial:', 'B16', 'piso_salarial'),
        ('A17', 'Contribuição Sindical:', 'B17', 'contribuicao_sindical'),
        ('A18', 'Benefícios:', 'B18', 'beneficios'),
        ('A19', 'Multa:', 'B19', 'multa'),
        ('A20', 'Jornada:', 'B20', 'jornada'),
        ('A21', 'Aviso Prévio:', 'B21', 'aviso_previo'),
        ('A22', 'Resumo Sintético:', 'B22', 'resumo_sintetico'),
    ]

    for label_cell, label_text, value_cell, col_name in campos:
        # Label
        ws_dash[label_cell] = label_text
        ws_dash[label_cell].font = Font(bold=True, size=10)
        ws_dash[label_cell].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        ws_dash[label_cell].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

        # Valor (usar fórmula PROCV para buscar dinamicamente)
        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = ws_dados.cell(row=1, column=col_idx).column_letter

        # Fórmula: =SE(B5="","",PROCV(VALOR(EXT.TEXTO(B5;1;ACHAR(" -";B5)-1));Dados!A:C;3;FALSO))
        # Simplificando: usar PROCV com MATCH
        formula = f'=IF(B5="","",VLOOKUP(VALUE(LEFT(B5,FIND(" -",B5)-1)),Dados!A:{col_letter},{col_idx},FALSE))'

        ws_dash[value_cell] = formula
        ws_dash[value_cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws_dash[value_cell].font = Font(size=10)
        ws_dash[value_cell].border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )

    # Ajustar largura da coluna B no dashboard
    ws_dash.column_dimensions['B'].width = 80
    ws_dash.column_dimensions['A'].width = 28

    # Ajustar altura das linhas
    for row in range(7, 23):
        ws_dash.row_dimensions[row].height = 40

    # Estilo de bordas para labels
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    for label_cell, _, value_cell, _ in campos:
        ws_dash[label_cell].border = thin_border

    # Esconder coluna H
    ws_dash.column_dimensions['H'].hidden = True

    # Instruções no rodapé
    ws_dash['A24'] = "Instruções:"
    ws_dash['A24'].font = Font(bold=True, size=10)
    ws_dash.merge_cells('A24:F24')
    ws_dash['A25'] = "• Para atualizar os dados, execute novamente o script extrair_cct.py"
    ws_dash['A25'].font = Font(size=9, color="666666")
    ws_dash.merge_cells('A25:F25')
    ws_dash['A26'] = "• Os dados são extraídos automaticamente dos PDFs da pasta 'convencoes'"
    ws_dash['A26'].font = Font(size=9, color="666666")
    ws_dash.merge_cells('A26:F26')

    # Salvar
    wb.save(ARQUIVO_SAIDA)
    print(f"\nPlanilha salva em: {ARQUIVO_SAIDA}")
    print(f"Total de convenções processadas: {len(resultados)}")


# =============================================================================
# MAIN
# =============================================================================

if __name__ == "__main__":
    print("=" * 60)
    print("EXTRATOR DE CONVENÇÕES COLETIVAS")
    print("=" * 60)
    print(f"Pasta de origem: {PASTA_PDFS}")
    print(f"Arquivo de saída: {ARQUIVO_SAIDA}")
    print("=" * 60)

    resultados = processar_todos_pdfs()

    if resultados:
        criar_excel(resultados)
    else:
        print("Nenhum PDF processado com sucesso.")

    print("\nConcluído!")
