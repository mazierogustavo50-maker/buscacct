#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Dashboard CCT - Versão Caprichada
Gera um dashboard profissional com:
- Aba Dados: formatação zebra, filtros, cores profissionais
- Aba Dashboard: layout aprimorado com dropdown
- Aba Resumo: estatísticas gerais
"""

import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from datetime import datetime

PASTA_DADOS = r"/mnt/c/Script/mediadorcct/dados"
ARQUIVO_ENTRADA = os.path.join(PASTA_DADOS, "dashboard_cct.xlsx")
ARQUIVO_SAIDA = os.path.join(PASTA_DADOS, "dashboard_cct_final.xlsx")


def criar_borda():
    return Border(
        left=Side(style='thin', color='B4B4B4'),
        right=Side(style='thin', color='B4B4B4'),
        top=Side(style='thin', color='B4B4B4'),
        bottom=Side(style='thin', color='B4B4B4')
    )


def criar_borda_header():
    return Border(
        left=Side(style='medium', color='1F4E78'),
        right=Side(style='medium', color='1F4E78'),
        top=Side(style='medium', color='1F4E78'),
        bottom=Side(style='medium', color='1F4E78')
    )


def ajustar_largura_colunas(ws, df):
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for cell in ws[col_letter][1:]:
            if cell.value:
                max_len = max(max_len, min(len(str(cell.value)), 80))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


def criar_aba_dados(wb, df):
    ws = wb.active
    ws.title = "Dados"

    # Cabeçalho
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")

    for c_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=c_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = criar_borda_header()

    ws.row_dimensions[1].height = 35

    # Dados com zebra
    fill_par = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    data_font = Font(name="Calibri", size=10)

    for r_idx, row in enumerate(df.itertuples(index=False), 2):
        fill = fill_par if r_idx % 2 == 0 else fill_impar
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = criar_borda()

    # Ajustar larguras
    ajustar_largura_colunas(ws, df)

    # Congelar painel e filtro
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    return ws


def criar_aba_dashboard(wb, df, ws_dados):
    ws = wb.create_sheet("Dashboard")

    # Título principal
    ws['A1'] = "DASHBOARD - CONVENÇÕES COLETIVAS DE TRABALHO"
    ws['A1'].font = Font(size=18, bold=True, color="1F4E78", name="Calibri")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:G1')
    ws.row_dimensions[1].height = 40

    # Subtítulo
    ws['A2'] = f"Atualizado em: {datetime.now().strftime('%d/%m/%Y %H:%M')} | Total de Convenções: {len(df)}"
    ws['A2'].font = Font(italic=True, color="666666", size=10, name="Calibri")
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 25

    # Linha separadora
    for col in range(1, 8):
        cell = ws.cell(row=3, column=col)
        cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws.row_dimensions[3].height = 3

    # Instrução
    ws['A5'] = "Selecione uma convenção no dropdown abaixo para visualizar os detalhes:"
    ws['A5'].font = Font(bold=True, size=11, name="Calibri", color="1F4E78")
    ws.merge_cells('A5:G5')
    ws.row_dimensions[5].height = 22

    # Label e dropdown
    ws['A7'] = "Convenção:"
    ws['A7'].font = Font(bold=True, size=11, name="Calibri", color="1F4E78")
    ws['A7'].alignment = Alignment(horizontal="right", vertical="center")
    ws['A7'].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws['A7'].border = criar_borda()
    ws.merge_cells('A7:B7')

    opcoes = [f"{row['id']} - {row['nome_arquivo']}" for _, row in df.iterrows()]
    col_oculta = 'I'
    for i, opcao in enumerate(opcoes, 1):
        ws[f'{col_oculta}{i}'] = opcao

    dv = DataValidation(
        type="list",
        formula1=f"${col_oculta}$1:${col_oculta}${len(opcoes)}",
        allow_blank=True
    )
    dv.error = 'Selecione uma convenção da lista'
    dv.errorTitle = 'Seleção inválida'
    ws.add_data_validation(dv)
    dv.add(ws['C7'])

    ws['C7'].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws['C7'].font = Font(size=10, name="Calibri")
    ws['C7'].border = criar_borda()
    ws.merge_cells('C7:G7')
    ws.row_dimensions[7].height = 30
    ws.column_dimensions['C'].width = 80

    # Campos do dashboard
    campos = [
        ('A9', 'Empregados (Parte):', 'C9', 'parte_empregados'),
        ('A10', 'Empregadores (Parte):', 'C10', 'parte_empregadores'),
        ('A11', 'CNPJ Empregados:', 'C11', 'cnpj_empregados'),
        ('A12', 'CNPJ Empregadores:', 'C12', 'cnpj_empregadores'),
        ('A13', 'Vigência:', 'C13', 'vigencia_fim'),
        ('A14', 'Data-Base:', 'C14', 'data_base'),
        ('A15', 'Reajuste:', 'C15', 'reajuste'),
        ('A16', 'Pisos Salariais:', 'C16', 'pisos_salariais'),
        ('A17', 'Contribuição Empregados:', 'C17', 'contribuicao_empregados'),
        ('A18', 'Contribuição Patronal:', 'C18', 'contribuicao_patronal'),
        ('A19', 'Benefícios:', 'C19', 'beneficios'),
        ('A20', 'Jornada:', 'C20', 'jornada'),
        ('A21', 'Aviso Prévio:', 'C21', 'aviso_previo'),
        ('A22', 'Multa:', 'C22', 'multa'),
    ]

    label_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    label_font = Font(bold=True, size=10, name="Calibri", color="1F4E78")
    value_font = Font(size=10, name="Calibri")

    for label_cell, label_text, value_cell, col_name in campos:
        row_num = int("".join(filter(str.isdigit, value_cell)))
        
        # Label em A e B
        ws[label_cell] = label_text
        ws[label_cell].font = label_font
        ws[label_cell].alignment = Alignment(horizontal="right", vertical="top", wrap_text=True)
        ws[label_cell].fill = label_fill
        ws[label_cell].border = criar_borda()
        ws.merge_cells(f'A{row_num}:B{row_num}')

        col_idx = df.columns.get_loc(col_name) + 1
        col_letter = get_column_letter(col_idx)
        formula = f'=IF(C7="","",VLOOKUP(VALUE(LEFT(C7,FIND(" -",C7)-1)),Dados!A:{col_letter},{col_idx},FALSE))'

        ws[value_cell] = formula
        ws[value_cell].alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws[value_cell].font = value_font
        ws[value_cell].border = criar_borda()
        ws.merge_cells(f'C{row_num}:G{row_num}')
        ws.row_dimensions[row_num].height = 55

    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 4

    # Rodapé
    ws['A24'] = "Instruções:"
    ws['A24'].font = Font(bold=True, size=10, name="Calibri", color="1F4E78")
    ws.merge_cells('A24:G24')

    ws['A25'] = "• Para atualizar os dados, execute: python extrair_cct_v2.py"
    ws['A25'].font = Font(size=9, color="666666", name="Calibri")
    ws.merge_cells('A25:G25')

    ws['A26'] = "• Os dados são extraídos automaticamente dos PDFs da pasta 'convencoes'"
    ws['A26'].font = Font(size=9, color="666666", name="Calibri")
    ws.merge_cells('A26:G26')

    ws.column_dimensions['I'].hidden = True

    return ws


def criar_aba_resumo(wb, df):
    ws = wb.create_sheet("Resumo")

    # Título
    ws['A1'] = "RESUMO ESTATÍSTICO - CCTs"
    ws['A1'].font = Font(size=16, bold=True, color="1F4E78", name="Calibri")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 35

    ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].font = Font(italic=True, color="666666", size=10, name="Calibri")
    ws.merge_cells('A2:D2')

    # Estatísticas
    stats = [
        ("Total de Convenções", len(df)),
        ("Com Vigência Definida", df['vigencia_inicio'].notna().sum()),
        ("Com Data-Base", df['data_base'].notna().sum()),
        ("Com Reajuste", df['reajuste'].notna().sum()),
        ("Com Pisos Salariais", df['pisos_salariais'].notna().sum()),
        ("Com Contribuição Empregados", df['contribuicao_empregados'].notna().sum()),
        ("Com Contribuição Patronal", df['contribuicao_patronal'].notna().sum()),
        ("Com Benefícios", (df['beneficios'] != 'Não especificado').sum()),
        ("Com Jornada Específica", df['jornada'].notna().sum()),
        ("Com Aviso Prévio", df['aviso_previo'].notna().sum()),
        ("Com Multa", df['multa'].notna().sum()),
    ]

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    ws['A4'] = "Métrica"
    ws['A4'].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    ws['A4'].fill = header_fill
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A4'].border = criar_borda()

    ws['B4'] = "Quantidade"
    ws['B4'].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    ws['B4'].fill = header_fill
    ws['B4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['B4'].border = criar_borda()

    ws['C4'] = "% do Total"
    ws['C4'].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
    ws['C4'].fill = header_fill
    ws['C4'].alignment = Alignment(horizontal="center", vertical="center")
    ws['C4'].border = criar_borda()

    fill_par = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    fill_impar = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for i, (label, value) in enumerate(stats, 5):
        fill = fill_par if i % 2 == 0 else fill_impar
        pct = (value / len(df) * 100) if len(df) > 0 else 0

        ws.cell(row=i, column=1, value=label).font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=1).fill = fill
        ws.cell(row=i, column=1).border = criar_borda()

        ws.cell(row=i, column=2, value=value).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=i, column=2).fill = fill
        ws.cell(row=i, column=2).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=2).border = criar_borda()

        ws.cell(row=i, column=3, value=f"{pct:.1f}%").font = Font(name="Calibri", size=10)
        ws.cell(row=i, column=3).fill = fill
        ws.cell(row=i, column=3).alignment = Alignment(horizontal="center")
        ws.cell(row=i, column=3).border = criar_borda()

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15

    # Top reajustes
    ws['E4'] = "TOP 10 - MAIORES REAJUSTES"
    ws['E4'].font = Font(size=12, bold=True, color="1F4E78", name="Calibri")
    ws['E4'].alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells('E4:G4')
    ws['E4'].fill = header_fill
    ws['E4'].font = Font(bold=True, color="FFFFFF", size=11, name="Calibri")

    ws['E5'] = "ID"
    ws['F5'] = "Arquivo"
    ws['G5'] = "Reajuste"
    for col in ['E', 'F', 'G']:
        ws[f'{col}5'].font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        ws[f'{col}5'].fill = header_fill
        ws[f'{col}5'].alignment = Alignment(horizontal="center", vertical="center")
        ws[f'{col}5'].border = criar_borda()

    df_reaj = df[df['reajuste'].notna()].copy()
    df_reaj['reajuste_num'] = df_reaj['reajuste'].str.replace('%', '').str.replace(',', '.').astype(float)
    df_reaj = df_reaj.sort_values('reajuste_num', ascending=False).head(10)

    for i, (_, row) in enumerate(df_reaj.iterrows(), 6):
        fill = fill_par if i % 2 == 0 else fill_impar
        ws.cell(row=i, column=5, value=row['id']).font = Font(name="Calibri", size=9)
        ws.cell(row=i, column=5).fill = fill
        ws.cell(row=i, column=5).border = criar_borda()
        ws.cell(row=i, column=5).alignment = Alignment(horizontal="center")

        ws.cell(row=i, column=6, value=row['nome_arquivo'][:60]).font = Font(name="Calibri", size=9)
        ws.cell(row=i, column=6).fill = fill
        ws.cell(row=i, column=6).border = criar_borda()

        ws.cell(row=i, column=7, value=row['reajuste']).font = Font(name="Calibri", size=9, bold=True)
        ws.cell(row=i, column=7).fill = fill
        ws.cell(row=i, column=7).border = criar_borda()
        ws.cell(row=i, column=7).alignment = Alignment(horizontal="center")

    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 50
    ws.column_dimensions['G'].width = 12

    return ws


def main():
    print("=" * 60)
    print("GERANDO DASHBOARD CCT - VERSÃO CAPRICHADA")
    print("=" * 60)

    df = pd.read_excel(ARQUIVO_ENTRADA, sheet_name='Dados')
    print(f"Total de registros: {len(df)}")

    wb = Workbook()

    # Aba Dados
    print("Criando aba 'Dados'...")
    ws_dados = criar_aba_dados(wb, df)

    # Aba Dashboard
    print("Criando aba 'Dashboard'...")
    criar_aba_dashboard(wb, df, ws_dados)

    # Aba Resumo
    print("Criando aba 'Resumo'...")
    criar_aba_resumo(wb, df)

    # Salvar
    wb.save(ARQUIVO_SAIDA)
    print(f"\nDashboard salvo em: {ARQUIVO_SAIDA}")
    print("=" * 60)


if __name__ == "__main__":
    main()
